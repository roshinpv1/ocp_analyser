import os
import re
import yaml
import json
import openpyxl
from openpyxl import load_workbook
from urllib.parse import urlparse
from typing import Dict, List, Optional, Union
from core.genflow import Node, BatchNode
from utils.crawl_github_files import crawl_github_files
from utils.call_llm import call_llm
from utils.crawl_local_files import crawl_local_files
import time
import traceback


# OpenShift Assessment Node
class OcpAssessmentNode(Node):
    def prep(self, shared):
        print("\nDEBUG: Starting OcpAssessmentNode prep")
        
        # Get Excel validation data from the shared state
        excel_validation = shared.get("excel_validation", {})
        excel_file = shared.get("excel_file")
        component_name = shared.get("project_name", "Unknown Component")
        output_dir = shared.get("output_dir", "analysis_output")
        
        # Extract Excel data to be used for OCP assessment
        excel_data = {
            "component_name": component_name,
            "validation_results": excel_validation,
            "excel_file_path": excel_file
        }
        
        # If we have component questions, include them
        if "component_questions" in excel_validation:
            excel_data["component_questions"] = excel_validation["component_questions"]
            
        # Add additional data from the Excel file if needed
        try:
            if excel_file and os.path.exists(excel_file):
                wb = load_workbook(excel_file, read_only=True)
                sheet = wb.active
                
                # Extract a more comprehensive dataset from Excel
                all_data = {}
                
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                    if len(row) >= 3 and row[1].value:  # Question in second column
                        question = str(row[1].value).strip() if row[1].value else ""
                        answer = str(row[2].value).strip() if row[2].value else ""
                        
                        if question:
                            all_data[question] = answer
                
                excel_data["all_form_data"] = all_data
        except Exception as e:
            print(f"Error extracting additional Excel data: {str(e)}")
            
        return excel_data, output_dir

    def exec(self, prep_res):
        excel_data, output_dir = prep_res
        print(f"\nPerforming OpenShift migration assessment for component: {excel_data.get('component_name', 'Unknown')}")
        
        # Define the system prompt for the OpenShift assessment
        system_prompt = """You are an OpenShift migration intake assessment agent, designed to evaluate if application components can be migrated to OpenShift.
Your primary functions are:
1. Parse intake forms submitted as Excel documents, extracting all relevant application metadata including:
- Application name, ID, and business criticality
- Current hosting environment details and configurations
- Application dependencies and integration points
- Current resource utilization metrics and patterns
2. Identify source platforms (TAS, TKGI, on-prem VM, traditional servers) and their specific characteristics
- Operating system details and version compatibility
- Middleware components and their OpenShift equivalence
- External service dependencies and communication patterns
- Current deployment and operational procedures
- Existing monitoring and logging mechanisms
3. Perform comprehensive validation of intake form data including:
- Required field completion with specific validation rules for each field
- Data consistency across related fields and dependencies
- Technical feasibility checks against OpenShift compatibility matrix
- Resource specification validation against OpenShift limits and quotas
- Identification of missing critical information with specific requests for completion
- List down all the missing information in tabuler format
4. Conduct detailed migration assessment considering the given data if the application component can be migrated to openshift platform or not.
5. Generate structured assessment reports that include:
- Migration feasibility score (High/Medium/Low) with numeric ratings (0-100)
- Detailed scoring breakdown across 10+ technical dimensions
- Identified migration blockers or concerns with severity classification
- Required architectural changes with component-level details
- Suggested migration strategy (lift-and-shift,
re-platform, or re-architect) with justification
- Risk assessment with likelihood and impact analysis
- Do not include Assessment Date in the final repost
6. Provide actionable recommendations to address migration challenges:
- Configuration changes required for OpenShift compatibility
- Data migration approaches for stateful components
- Network configuration recommendations for service communication
- Security posture improvements for containerized environment
- Required application code changes with examples where applicable
- Testing strategy recommendations for migration validation
7. Identify potential optimization opportunities for containerized workloads:
- Resource right-sizing recommendations based on utilization patterns
- Horizontal vs vertical scaling
recommendations
- Service mesh adoption benefits for the specific application
The output must be formatted as a structured assessment report with clear sections, using tables where appropriate for comparative data. Do not extrapolate any data - be specific to display if the OpenShift assessment is
successful or not successful with required scoring (0-100 scale with detailed rubric) and specific recommendations/reasons. Include an executive summary with go/no-go recommendation,
followed by detailed technical sections.
Each finding must include evidence from the intake data and reference to relevant OpenShift constraints or requirements.
The output should be formatted well to view in html documents."""
        
        # Instruct the LLM to return content suitable for HTML embedding, not a full HTML document
        user_content = f"""Perform the OCP intake assessment for the following data and generate the assessment report content (HTML compatible):

{json.dumps(excel_data, indent=2)}

DO NOT return a full HTML document with <html>, <head>, or <body> tags.
Instead, return only the content that would go inside the <body> tag, with proper HTML formatting and styling.
Use <div>, <h1>, <h2>, <p>, <table>, etc. elements to structure your report.
"""
        
        # Combine into a complete prompt
        prompt = f"""
System: {system_prompt}

User: {user_content}

Generate a well-structured OpenShift migration assessment report with HTML formatting (excluding html/head/body tags).
Include detailed scoring with a clear go/no-go recommendation and use tables for comparison data.
"""
        
        try:
            # Call LLM with the prompt
            print("Calling LLM for OpenShift migration assessment...")
            response = call_llm(prompt, use_cache=(False))  # Always generate fresh assessment
            
            # Use the same styling as the main analysis report for consistency
            css_styles = """
                body {
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                    line-height: 1.4;
                    margin: 0;
                    padding: 15px;
                    color: #333;
                    max-width: 1200px;
                    margin: 0 auto;
                }
                h1 { font-size: 1.5em; margin: 0 0 15px 0; padding-bottom: 5px; border-bottom: 2px solid #eee; }
                h2 { font-size: 1.2em; margin: 15px 0 10px 0; color: #444; padding-top: 10px; border-top: 1px solid #eee; }
                h2:first-of-type { border-top: none; }
                h3 { font-size: 1.1em; margin: 10px 0 5px 0; color: #555; }
                .section {
                    margin: 10px 0;
                    padding: 10px;
                    background: #f8f9fa;
                    border-radius: 4px;
                }
                .component-table {
                    width: 100%;
                    border-collapse: collapse;
                    margin: 15px 0;
                }
                .component-table th, .component-table td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                .component-table th {
                    background-color: #f2f2f2;
                    font-weight: bold;
                }
                .component-table tr:nth-child(even) {
                    background-color: #f9f9f9;
                }
                .finding-list {
                    display: grid;
                    grid-template-columns: repeat(auto-fill, minmax(400px, 1fr));
                    gap: 10px;
                    margin: 10px 0;
                }
                .finding {
                    background: white;
                    padding: 10px;
                    border-radius: 4px;
                    border: 1px solid #eee;
                }
                .action-item {
                    background: #f8f9fa;
                    padding: 10px 15px;
                    margin: 10px 0;
                    border-radius: 4px;
                    border-left: 5px solid #007bff;
                }
                .action-item.critical {
                    border-left-color: #dc3545;
                }
                .action-item.high {
                    border-left-color: #fd7e14;
                }
                .action-item.medium {
                    border-left-color: #ffc107;
                }
                .action-item.low {
                    border-left-color: #28a745;
                }
                .severity-high { color: #dc3545; }
                .severity-medium { color: #fd7e14; }
                .severity-low { color: #28a745; }
                .severity-badge {
                    display: inline-block;
                    padding: 2px 6px;
                    border-radius: 3px;
                    font-size: 0.9em;
                    font-weight: 500;
                    margin-right: 8px;
                }
                .severity-high .severity-badge { background: #dc3545; color: white; }
                .severity-medium .severity-badge { background: #fd7e14; color: white; }
                .severity-low .severity-badge { background: #28a745; color: white; }
                .validation-status {
                    display: inline-block;
                    padding: 5px 10px;
                    border-radius: 4px;
                    font-weight: bold;
                }
                .validation-complete { background: #d4edda; color: #155724; }
                .validation-incomplete { background: #f8d7da; color: #721c24; }
                .component-yes {
                    color: #28a745;
                    font-weight: bold;
                }
                .component-no {
                    color: #dc3545;
                }
                .match {
                    color: #28a745;
                    font-weight: bold;
                }
                .mismatch {
                    color: #fd7e14;
                    font-weight: bold;
                }
                .toc {
                    background: #f8f9fa;
                    padding: 10px 15px;
                    border-radius: 4px;
                    margin: 15px 0;
                }
                .toc ul {
                    margin: 5px 0;
                    padding-left: 20px;
                }
                .executive-summary {
                    background: #e9f7fd;
                    border-radius: 4px;
                    padding: 10px 15px;
                    margin: 15px 0;
                    border-left: 5px solid #17a2b8;
                }
                pre, code {
                    background: #f8f9fa;
                    padding: 2px 4px;
                    border-radius: 3px;
                    font-family: monospace;
                    font-size: 0.9em;
                    overflow-x: auto;
                }
                .summary {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 10px;
                    margin: 10px 0;
                }
                .summary-item {
                    background: white;
                    padding: 10px;
                    border-radius: 4px;
                    border: 1px solid #eee;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin: 15px 0;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                th {
                    background-color: #f2f2f2;
                    font-weight: bold;
                }
                tr:nth-child(even) {
                    background-color: #f9f9f9;
                }
                .score-high { color: #28a745; font-weight: bold; }
                .score-medium { color: #fd7e14; font-weight: bold; }
                .score-low { color: #dc3545; font-weight: bold; }
                .assessment-score {
                    font-size: 24px;
                    font-weight: bold;
                    text-align: center;
                    margin: 10px 0;
                    padding: 10px;
                    border-radius: 4px;
                }
                .go-recommendation {
                    font-size: 18px;
                    font-weight: bold;
                    text-align: center;
                    margin: 10px 0;
                    padding: 10px;
                    border-radius: 4px;
                }
                .go { background-color: #d4edda; color: #155724; }
                .no-go { background-color: #f8d7da; color: #721c24; }
                .conditional-go { background-color: #fff3cd; color: #856404; }
                @media print {
                    body { padding: 10px; }
                    .section { break-inside: avoid; }
                }
            """
            
            # Wrap the response in a properly structured HTML document
            report_html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>OpenShift Migration Assessment Report - {excel_data.get('component_name', 'Unknown Component')}</title>
    <style>
{css_styles}
    </style>
</head>
<body>
    <h1>OpenShift Migration Assessment Report - {excel_data.get('component_name', 'Unknown Component')}</h1>
    
    <div class="section">
{response}
    </div>
</body>
</html>
"""
            
            # Save the report
            os.makedirs(output_dir, exist_ok=True)
            ocp_report_path = os.path.join(output_dir, "ocp_assessment_report.html")
            
            with open(ocp_report_path, "w", encoding="utf-8") as f:
                f.write(report_html)
                
            print(f"OpenShift migration assessment completed. Report saved to: {ocp_report_path}")
            
            # Also save as Markdown for consistency
            try:
                # Basic conversion from HTML to Markdown
                markdown_content = response.replace("<h1>", "# ").replace("</h1>", "\n\n")
                markdown_content = markdown_content.replace("<h2>", "## ").replace("</h2>", "\n\n")
                markdown_content = markdown_content.replace("<h3>", "### ").replace("</h3>", "\n\n")
                markdown_content = markdown_content.replace("<p>", "").replace("</p>", "\n\n")
                markdown_content = markdown_content.replace("<br>", "\n")
                markdown_content = markdown_content.replace("<strong>", "**").replace("</strong>", "**")
                markdown_content = markdown_content.replace("<em>", "*").replace("</em>", "*")
                
                # Save markdown version
                md_path = os.path.join(output_dir, "ocp_assessment_report.md")
                with open(md_path, "w", encoding="utf-8") as f:
                    f.write(f"# OpenShift Migration Assessment Report - {excel_data.get('component_name', 'Unknown Component')}\n\n")
                    f.write(markdown_content)
                
                print(f"Markdown version saved to: {md_path}")
                
            except Exception as e:
                print(f"Error saving Markdown version: {str(e)}")
            
            return {
                "html_report": ocp_report_path,
                "component_name": excel_data.get('component_name', 'Unknown Component')
            }
            
        except Exception as e:
            print(f"Error generating OpenShift migration assessment: {str(e)}")
            traceback.print_exc()
            
            # Return a basic error report using the same styling
            error_report_path = os.path.join(output_dir, "ocp_assessment_error.html")
            with open(error_report_path, "w", encoding="utf-8") as f:
                f.write(f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>OpenShift Migration Assessment Error</title>
    <style>
{css_styles}
    </style>
</head>
<body>
    <h1>OpenShift Migration Assessment Error</h1>
    <div class="section" style="background: #f8d7da; color: #721c24;">
        <p>An error occurred while generating the OpenShift migration assessment report:</p>
        <pre>{str(e)}</pre>
    </div>
    <p>Please check the Excel file and try again.</p>
</body>
</html>
""")
            
            return {
                "html_report": error_report_path,
                "component_name": excel_data.get('component_name', 'Unknown Component'),
                "error": str(e)
            }

    def post(self, shared, prep_res, exec_res):
        if not exec_res:
            print("No OpenShift assessment results available.")
            return "default"
            
        # Store the assessment report path in shared state
        shared["ocp_assessment_report"] = exec_res.get("html_report")
        
        # Print the report path
        print(f"\nOpenShift Migration Assessment completed for {exec_res.get('component_name')}!")
        print(f"Report saved to: {exec_res.get('html_report')}")
        
        if "error" in exec_res:
            print(f"Note: An error occurred during assessment: {exec_res['error']}")
        
        # Continue the normal flow
        return "default"


# Helper to get content for specific file indices
def get_content_for_indices(files_data, indices):
    content_map = {}
    for i in indices:
        if 0 <= i < len(files_data):
            path, content = files_data[i]
            content_map[f"{i} # {path}"] = (
                content  # Use index + path as key for context
            )
    return content_map


class FetchRepo(Node):
    def prep(self, shared):
        repo_url = shared.get("repo_url")
        local_dir = shared.get("local_dir")
        project_name = shared.get("project_name")

        if not project_name:
            # Basic name derivation from URL or directory
            if repo_url:
                project_name = repo_url.split("/")[-1].replace(".git", "")
            else:
                project_name = os.path.basename(os.path.abspath(local_dir))
            shared["project_name"] = project_name

        # Get file patterns directly from shared
        include_patterns = shared["include_patterns"]
        exclude_patterns = shared["exclude_patterns"]
        max_file_size = shared["max_file_size"]

        return {
            "repo_url": repo_url,
            "local_dir": local_dir,
            "token": shared.get("github_token"),
            "include_patterns": include_patterns,
            "exclude_patterns": exclude_patterns,
            "max_file_size": max_file_size,
            "use_relative_paths": True,
            "shallow_clone": True,  # Enable shallow clone by default
            "clone_timeout": 300,   # 5 minutes timeout
            "max_depth": 1         # Only clone the latest commit
        }

    def exec(self, prep_res):
        if prep_res["repo_url"]:
            print(f"Crawling repository: {prep_res['repo_url']}...")
            max_retries = 3
            retry_delay = 5  # seconds
            
            for attempt in range(max_retries):
                try:
                    result = crawl_github_files(
                        repo_url=prep_res["repo_url"],
                        token=prep_res["token"],
                        include_patterns=prep_res["include_patterns"],
                        exclude_patterns=prep_res["exclude_patterns"],
                        max_file_size=prep_res["max_file_size"],
                        use_relative_paths=prep_res["use_relative_paths"],
                        shallow_clone=prep_res["shallow_clone"],
                        clone_timeout=prep_res["clone_timeout"],
                        max_depth=prep_res["max_depth"]
                    )
                    break  # If successful, break the retry loop
                except TimeoutError as e:
                    if attempt < max_retries - 1:
                        print(f"\nAttempt {attempt + 1} failed: {str(e)}")
                        print(f"Retrying in {retry_delay} seconds...")
                        time.sleep(retry_delay)
                        retry_delay *= 2  # Exponential backoff
                    else:
                        print("\nAll retry attempts failed. Last error:")
                        raise ValueError(f"Failed to clone repository after {max_retries} attempts: {str(e)}")
                except Exception as e:
                    if attempt < max_retries - 1:
                        print(f"\nAttempt {attempt + 1} failed: {str(e)}")
                        print(f"Retrying in {retry_delay} seconds...")
                        time.sleep(retry_delay)
                        retry_delay *= 2  # Exponential backoff
                    else:
                        print("\nAll retry attempts failed. Last error:")
                        raise ValueError(f"Failed to clone repository after {max_retries} attempts: {str(e)}")
        elif prep_res["local_dir"]:
            print(f"Crawling directory: {prep_res['local_dir']}...")
            result = crawl_local_files(
                directory=prep_res["local_dir"],
                include_patterns=prep_res["include_patterns"],
                exclude_patterns=prep_res["exclude_patterns"],
                max_file_size=prep_res["max_file_size"],
                use_relative_paths=prep_res["use_relative_paths"]
            )
        else:
            # Neither repo_url nor local_dir is specified
            print("Error: No repository URL or local directory specified.")
            raise ValueError("No repository URL or local directory specified. Cannot proceed with code analysis.")

        # Convert dict to list of tuples: [(path, content), ...]
        files_list = list(result.get("files", {}).items())
        
        # Check if any files were found
        if len(files_list) == 0:
            # Get the source for better error message
            source = prep_res["repo_url"] or prep_res["local_dir"]
            include_patterns = prep_res["include_patterns"]
            exclude_patterns = prep_res["exclude_patterns"]
            
            error_msg = f"No files found in {source} matching the patterns:\n"
            error_msg += f"Include patterns: {include_patterns}\n"
            error_msg += f"Exclude patterns: {exclude_patterns}\n\n"
            error_msg += "Please check:\n"
            error_msg += "1. The repository/directory path is correct\n"
            error_msg += "2. The include/exclude patterns match your files\n"
            error_msg += "3. The repository is accessible (if using GitHub)\n"
            error_msg += "4. The file extensions in include patterns match your codebase"
            
            raise ValueError(error_msg)
            
        print(f"Fetched {len(files_list)} files.")
        return files_list

    def post(self, shared, prep_res, exec_res):
        shared["files"] = exec_res  # List of (path, content) tuples


class IdentifyAbstractions1(Node):
    def prep(self, shared):
        files_data = shared["files"]
        project_name = shared["project_name"]  # Get project name
        language = shared.get("language", "english")  # Get language
        use_cache = shared.get("use_cache", True)  # Get use_cache flag, default to True
        max_abstraction_num = shared.get("max_abstraction_num", 10)  # Get max_abstraction_num, default to 10

        # Helper to create context from files, respecting limits (basic example)
        def create_llm_context(files_data):
            context = ""
            file_info = []  # Store tuples of (index, path)
            for i, (path, content) in enumerate(files_data):
                entry = f"--- File Index {i}: {path} ---\n{content}\n\n"
                context += entry
                file_info.append((i, path))

            return context, file_info  # file_info is list of (index, path)

        context, file_info = create_llm_context(files_data)
        # Format file info for the prompt (comment is just a hint for LLM)
        file_listing_for_prompt = "\n".join(
            [f"- {idx} # {path}" for idx, path in file_info]
        )
        return (
            context,
            file_listing_for_prompt,
            len(files_data),
            project_name,
            language,
            use_cache,
            max_abstraction_num,
        )  # Return all parameters

    def exec(self, prep_res):
        (
            context,
            file_listing_for_prompt,
            file_count,
            project_name,
            language,
            use_cache,
            max_abstraction_num,
        ) = prep_res  # Unpack all parameters
        print(f"Identifying abstractions using LLM...")

        # Add language instruction and hints only if not English
        language_instruction = ""
        name_lang_hint = ""
        desc_lang_hint = ""
        if language.lower() != "english":
            language_instruction = f"IMPORTANT: Generate the `name` and `description` for each abstraction in **{language.capitalize()}** language. Do NOT use English for these fields.\n\n"
            # Keep specific hints here as name/description are primary targets
            name_lang_hint = f" (value in {language.capitalize()})"
            desc_lang_hint = f" (value in {language.capitalize()})"

        prompt = f"""
For the project `{project_name}`:

Codebase Context:
{context}

{language_instruction}Analyze the codebase context.
Identify the top 5-{max_abstraction_num} core most important abstractions to help those new to the codebase.

For each abstraction, provide:
1. A concise `name`{name_lang_hint}.
2. A beginner-friendly `description` explaining what it is with a simple analogy, in around 100 words{desc_lang_hint}.
3. A list of relevant `file_indices` (integers) using the format `idx # path/comment`.

List of file indices and paths present in the context:
{file_listing_for_prompt}

Format the output as a YAML list of dictionaries:

```yaml
- name: |
    Query Processing{name_lang_hint}
  description: |
    Explains what the abstraction does.
    It's like a central dispatcher routing requests.{desc_lang_hint}
  file_indices:
    - 0 # path/to/file1.py
    - 3 # path/to/related.py
- name: |
    Query Optimization{name_lang_hint}
  description: |
    Another core concept, similar to a blueprint for objects.{desc_lang_hint}
  file_indices:
    - 5 # path/to/another.js
# ... up to {max_abstraction_num} abstractions
```"""
        response = call_llm(prompt, use_cache=(use_cache and self.cur_retry == 0))  # Use cache only if enabled and not retrying

        # --- Validation ---
        yaml_str = response.strip().split("```yaml")[1].split("```")[0].strip()
        abstractions = yaml.safe_load(yaml_str)

        if not isinstance(abstractions, list):
            raise ValueError("LLM Output is not a list")

        validated_abstractions = []
        for item in abstractions:
            if not isinstance(item, dict) or not all(
                k in item for k in ["name", "description", "file_indices"]
            ):
                raise ValueError(f"Missing keys in abstraction item: {item}")
            if not isinstance(item["name"], str):
                raise ValueError(f"Name is not a string in item: {item}")
            if not isinstance(item["description"], str):
                raise ValueError(f"Description is not a string in item: {item}")
            if not isinstance(item["file_indices"], list):
                raise ValueError(f"file_indices is not a list in item: {item}")

            # Validate indices
            validated_indices = []
            for idx_entry in item["file_indices"]:
                try:
                    if isinstance(idx_entry, int):
                        idx = idx_entry
                    elif isinstance(idx_entry, str) and "#" in idx_entry:
                        idx = int(idx_entry.split("#")[0].strip())
                    else:
                        idx = int(str(idx_entry).strip())

                    if not (0 <= idx < file_count):
                        raise ValueError(
                            f"Invalid file index {idx} found in item {item['name']}. Max index is {file_count - 1}."
                        )
                    validated_indices.append(idx)
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Could not parse index from entry: {idx_entry} in item {item['name']}"
                    )

            item["files"] = sorted(list(set(validated_indices)))
            # Store only the required fields
            validated_abstractions.append(
                {
                    "name": item["name"],  # Potentially translated name
                    "description": item[
                        "description"
                    ],  # Potentially translated description
                    "files": item["files"],
                }
            )

        print(f"Identified {len(validated_abstractions)} abstractions.")
        return validated_abstractions

    def post(self, shared, prep_res, exec_res):
        shared["abstractions"] = (
            exec_res  # List of {"name": str, "description": str, "files": [int]}
        )


class AnalyzeRelationships1(Node):
    def prep(self, shared):
        abstractions = shared[
            "abstractions"
        ]  # Now contains 'files' list of indices, name/description potentially translated
        files_data = shared["files"]
        project_name = shared["project_name"]  # Get project name
        language = shared.get("language", "english")  # Get language
        use_cache = shared.get("use_cache", True)  # Get use_cache flag, default to True

        # Get the actual number of abstractions directly
        num_abstractions = len(abstractions)

        # Create context with abstraction names, indices, descriptions, and relevant file snippets
        context = "Identified Abstractions:\\n"
        all_relevant_indices = set()
        abstraction_info_for_prompt = []
        for i, abstr in enumerate(abstractions):
            # Use 'files' which contains indices directly
            file_indices_str = ", ".join(map(str, abstr["files"]))
            # Abstraction name and description might be translated already
            info_line = f"- Index {i}: {abstr['name']} (Relevant file indices: [{file_indices_str}])\\n  Description: {abstr['description']}"
            context += info_line + "\\n"
            abstraction_info_for_prompt.append(
                f"{i} # {abstr['name']}"
            )  # Use potentially translated name here too
            all_relevant_indices.update(abstr["files"])

        context += "\\nRelevant File Snippets (Referenced by Index and Path):\\n"
        # Get content for relevant files using helper
        relevant_files_content_map = get_content_for_indices(
            files_data, sorted(list(all_relevant_indices))
        )
        # Format file content for context
        file_context_str = "\\n\\n".join(
            f"--- File: {idx_path} ---\\n{content}"
            for idx_path, content in relevant_files_content_map.items()
        )
        context += file_context_str

        return (
            context,
            "\n".join(abstraction_info_for_prompt),
            num_abstractions, # Pass the actual count
            project_name,
            language,
            use_cache,
        )  # Return use_cache

    def exec(self, prep_res):
        (
            context,
            abstraction_listing,
            num_abstractions, # Receive the actual count
            project_name,
            language,
            use_cache,
         ) = prep_res  # Unpack use_cache
        print(f"Analyzing relationships using LLM...")

        # Add language instruction and hints only if not English
        language_instruction = ""
        lang_hint = ""
        list_lang_note = ""
        if language.lower() != "english":
            language_instruction = f"IMPORTANT: Generate the `summary` and relationship `label` fields in **{language.capitalize()}** language. Do NOT use English for these fields.\n\n"
            lang_hint = f" (in {language.capitalize()})"
            list_lang_note = f" (Names might be in {language.capitalize()})"  # Note for the input list

        prompt = f"""
Based on the following abstractions and relevant code snippets from the project `{project_name}`:

List of Abstraction Indices and Names{list_lang_note}:
{abstraction_listing}

Context (Abstractions, Descriptions, Code):
{context}

{language_instruction}Please provide:
1. A high-level `summary` of the project's main purpose and functionality in a few beginner-friendly sentences{lang_hint}. Use markdown formatting with **bold** and *italic* text to highlight important concepts.
2. A list (`relationships`) describing the key interactions between these abstractions. For each relationship, specify:
    - `from_abstraction`: Index of the source abstraction (e.g., `0 # AbstractionName1`)
    - `to_abstraction`: Index of the target abstraction (e.g., `1 # AbstractionName2`)
    - `label`: A brief label for the interaction **in just a few words**{lang_hint} (e.g., "Manages", "Inherits", "Uses").
    Ideally the relationship should be backed by one abstraction calling or passing parameters to another.
    Simplify the relationship and exclude those non-important ones.

IMPORTANT: Make sure EVERY abstraction is involved in at least ONE relationship (either as source or target). Each abstraction index must appear at least once across all relationships.

Format the output as YAML:

```yaml
summary: |
  A brief, simple explanation of the project{lang_hint}.
  Can span multiple lines with **bold** and *italic* for emphasis.
relationships:
  - from_abstraction: 0 # AbstractionName1
    to_abstraction: 1 # AbstractionName2
    label: "Manages"{lang_hint}
  - from_abstraction: 2 # AbstractionName3
    to_abstraction: 0 # AbstractionName1
    label: "Provides config"{lang_hint}
  # ... other relationships
```

Now, provide the YAML output:
"""
        response = call_llm(prompt, use_cache=(use_cache and self.cur_retry == 0)) # Use cache only if enabled and not retrying

        # --- Validation ---
        yaml_str = response.strip().split("```yaml")[1].split("```")[0].strip()
        relationships_data = yaml.safe_load(yaml_str)

        if not isinstance(relationships_data, dict) or not all(
            k in relationships_data for k in ["summary", "relationships"]
        ):
            raise ValueError(
                "LLM output is not a dict or missing keys ('summary', 'relationships')"
            )
        if not isinstance(relationships_data["summary"], str):
            raise ValueError("summary is not a string")
        if not isinstance(relationships_data["relationships"], list):
            raise ValueError("relationships is not a list")

        # Validate relationships structure
        validated_relationships = []
        for rel in relationships_data["relationships"]:
            # Check for 'label' key
            if not isinstance(rel, dict) or not all(
                k in rel for k in ["from_abstraction", "to_abstraction", "label"]
            ):
                raise ValueError(
                    f"Missing keys (expected from_abstraction, to_abstraction, label) in relationship item: {rel}"
                )
            # Validate 'label' is a string
            if not isinstance(rel["label"], str):
                raise ValueError(f"Relationship label is not a string: {rel}")

            # Validate indices
            try:
                from_idx = int(str(rel["from_abstraction"]).split("#")[0].strip())
                to_idx = int(str(rel["to_abstraction"]).split("#")[0].strip())
                if not (
                    0 <= from_idx < num_abstractions and 0 <= to_idx < num_abstractions
                ):
                    raise ValueError(
                        f"Invalid index in relationship: from={from_idx}, to={to_idx}. Max index is {num_abstractions-1}."
                    )
                validated_relationships.append(
                    {
                        "from": from_idx,
                        "to": to_idx,
                        "label": rel["label"],  # Potentially translated label
                    }
                )
            except (ValueError, TypeError):
                raise ValueError(f"Could not parse indices from relationship: {rel}")

        print("Generated project summary and relationship details.")
        return {
            "summary": relationships_data["summary"],  # Potentially translated summary
            "details": validated_relationships,  # Store validated, index-based relationships with potentially translated labels
        }

    def post(self, shared, prep_res, exec_res):
        # Structure is now {"summary": str, "details": [{"from": int, "to": int, "label": str}]}
        # Summary and label might be translated
        shared["relationships"] = exec_res


class OrderChapters1(Node):
    def prep(self, shared):
        abstractions = shared["abstractions"]  # Name/description might be translated
        relationships = shared["relationships"]  # Summary/label might be translated
        project_name = shared["project_name"]  # Get project name
        language = shared.get("language", "english")  # Get language
        use_cache = shared.get("use_cache", True)  # Get use_cache flag, default to True

        # Prepare context for the LLM
        abstraction_info_for_prompt = []
        for i, a in enumerate(abstractions):
            abstraction_info_for_prompt.append(
                f"- {i} # {a['name']}"
            )  # Use potentially translated name
        abstraction_listing = "\n".join(abstraction_info_for_prompt)

        # Use potentially translated summary and labels
        summary_note = ""
        if language.lower() != "english":
            summary_note = (
                f" (Note: Project Summary might be in {language.capitalize()})"
            )

        context = f"Project Summary{summary_note}:\n{relationships['summary']}\n\n"
        context += "Relationships (Indices refer to abstractions above):\n"
        for rel in relationships["details"]:
            from_name = abstractions[rel["from"]]["name"]
            to_name = abstractions[rel["to"]]["name"]
            # Use potentially translated 'label'
            context += f"- From {rel['from']} ({from_name}) to {rel['to']} ({to_name}): {rel['label']}\n"  # Label might be translated

        list_lang_note = ""
        if language.lower() != "english":
            list_lang_note = f" (Names might be in {language.capitalize()})"

        return (
            abstraction_listing,
            context,
            len(abstractions),
            project_name,
            list_lang_note,
            use_cache,
        )  # Return use_cache

    def exec(self, prep_res):
        (
            abstraction_listing,
            context,
            num_abstractions,
            project_name,
            list_lang_note,
            use_cache,
        ) = prep_res  # Unpack use_cache
        print("Determining chapter order using LLM...")
        # No language variation needed here in prompt instructions, just ordering based on structure
        # The input names might be translated, hence the note.
        prompt = f"""
Given the following project abstractions and their relationships for the project ```` {project_name} ````:

Abstractions (Index # Name){list_lang_note}:
{abstraction_listing}

Context about relationships and project summary:
{context}

If you are going to make a tutorial for ```` {project_name} ````, what is the best order to explain these abstractions, from first to last?
Ideally, first explain those that are the most important or foundational, perhaps user-facing concepts or entry points. Then move to more detailed, lower-level implementation details or supporting concepts.

Output the ordered list of abstraction indices, including the name in a comment for clarity. Use the format `idx # AbstractionName`.

```yaml
- 2 # FoundationalConcept
- 0 # CoreClassA
- 1 # CoreClassB (uses CoreClassA)
- ...
```

Now, provide the YAML output:
"""
        response = call_llm(prompt, use_cache=(use_cache and self.cur_retry == 0)) # Use cache only if enabled and not retrying

        # --- Validation ---
        yaml_str = response.strip().split("```yaml")[1].split("```")[0].strip()
        ordered_indices_raw = yaml.safe_load(yaml_str)

        if not isinstance(ordered_indices_raw, list):
            raise ValueError("LLM output is not a list")

        ordered_indices = []
        seen_indices = set()
        for entry in ordered_indices_raw:
            try:
                if isinstance(entry, int):
                    idx = entry
                elif isinstance(entry, str) and "#" in entry:
                    idx = int(entry.split("#")[0].strip())
                else:
                    idx = int(str(entry).strip())

                if not (0 <= idx < num_abstractions):
                    raise ValueError(
                        f"Invalid index {idx} in ordered list. Max index is {num_abstractions-1}."
                    )
                if idx in seen_indices:
                    raise ValueError(f"Duplicate index {idx} found in ordered list.")
                ordered_indices.append(idx)
                seen_indices.add(idx)

            except (ValueError, TypeError):
                raise ValueError(
                    f"Could not parse index from ordered list entry: {entry}"
                )

        # Check if all abstractions are included
        if len(ordered_indices) != num_abstractions:
            raise ValueError(
                f"Ordered list length ({len(ordered_indices)}) does not match number of abstractions ({num_abstractions}). Missing indices: {set(range(num_abstractions)) - seen_indices}"
            )

        print(f"Determined chapter order (indices): {ordered_indices}")
        return ordered_indices  # Return the list of indices

    def post(self, shared, prep_res, exec_res):
        # exec_res is already the list of ordered indices
        shared["chapter_order"] = exec_res  # List of indices


class WriteChapters1(BatchNode):
    def prep(self, shared):
        chapter_order = shared["chapter_order"]  # List of indices
        abstractions = shared[
            "abstractions"
        ]  # List of {"name": str, "description": str, "files": [int]}
        files_data = shared["files"]  # List of (path, content) tuples
        project_name = shared["project_name"]
        language = shared.get("language", "english")
        use_cache = shared.get("use_cache", True)  # Get use_cache flag, default to True

        # Get already written chapters to provide context
        # We store them temporarily during the batch run, not in shared memory yet
        # The 'previous_chapters_summary' will be built progressively in the exec context
        self.chapters_written_so_far = (
            []
        )  # Use instance variable for temporary storage across exec calls

        # Create a complete list of all chapters
        all_chapters = []
        chapter_filenames = {}  # Store chapter filename mapping for linking
        for i, abstraction_index in enumerate(chapter_order):
            if 0 <= abstraction_index < len(abstractions):
                chapter_num = i + 1
                chapter_name = abstractions[abstraction_index][
                    "name"
                ]  # Potentially translated name
                # Create safe filename (from potentially translated name)
                safe_name = "".join(
                    c if c.isalnum() else "_" for c in chapter_name
                ).lower()
                filename = f"{i+1:02d}_{safe_name}.md"
                # Format with link (using potentially translated name)
                all_chapters.append(f"{chapter_num}. [{chapter_name}]({filename})")
                # Store mapping of chapter index to filename for linking
                chapter_filenames[abstraction_index] = {
                    "num": chapter_num,
                    "name": chapter_name,
                    "filename": filename,
                }

        # Create a formatted string with all chapters
        full_chapter_listing = "\n".join(all_chapters)

        items_to_process = []
        for i, abstraction_index in enumerate(chapter_order):
            if 0 <= abstraction_index < len(abstractions):
                abstraction_details = abstractions[
                    abstraction_index
                ]  # Contains potentially translated name/desc
                # Use 'files' (list of indices) directly
                related_file_indices = abstraction_details.get("files", [])
                # Get content using helper, passing indices
                related_files_content_map = get_content_for_indices(
                    files_data, related_file_indices
                )

                # Get previous chapter info for transitions (uses potentially translated name)
                prev_chapter = None
                if i > 0:
                    prev_idx = chapter_order[i - 1]
                    prev_chapter = chapter_filenames[prev_idx]

                # Get next chapter info for transitions (uses potentially translated name)
                next_chapter = None
                if i < len(chapter_order) - 1:
                    next_idx = chapter_order[i + 1]
                    next_chapter = chapter_filenames[next_idx]

                items_to_process.append(
                    {
                        "chapter_num": i + 1,
                        "abstraction_index": abstraction_index,
                        "abstraction_details": abstraction_details,  # Has potentially translated name/desc
                        "related_files_content_map": related_files_content_map,
                        "project_name": shared["project_name"],  # Add project name
                        "full_chapter_listing": full_chapter_listing,  # Add the full chapter listing (uses potentially translated names)
                        "chapter_filenames": chapter_filenames,  # Add chapter filenames mapping (uses potentially translated names)
                        "prev_chapter": prev_chapter,  # Add previous chapter info (uses potentially translated name)
                        "next_chapter": next_chapter,  # Add next chapter info (uses potentially translated name)
                        "language": language,  # Add language for multi-language support
                        "use_cache": use_cache, # Pass use_cache flag
                        # previous_chapters_summary will be added dynamically in exec
                    }
                )
            else:
                print(
                    f"Warning: Invalid abstraction index {abstraction_index} in chapter_order. Skipping."
                )

        print(f"Preparing to write {len(items_to_process)} chapters...")
        return items_to_process  # Iterable for BatchNode

    def exec(self, item):
        # This runs for each item prepared above
        abstraction_name = item["abstraction_details"][
            "name"
        ]  # Potentially translated name
        abstraction_description = item["abstraction_details"][
            "description"
        ]  # Potentially translated description
        chapter_num = item["chapter_num"]
        project_name = item.get("project_name")
        language = item.get("language", "english")
        use_cache = item.get("use_cache", True) # Read use_cache from item
        print(f"Writing chapter {chapter_num} for: {abstraction_name} using LLM...")

        # Prepare file context string from the map
        file_context_str = "\n\n".join(
            f"--- File: {idx_path.split('# ')[1] if '# ' in idx_path else idx_path} ---\n{content}"
            for idx_path, content in item["related_files_content_map"].items()
        )

        # Get summary of chapters written *before* this one
        # Use the temporary instance variable
        previous_chapters_summary = "\n---\n".join(self.chapters_written_so_far)

        # Add language instruction and context notes only if not English
        language_instruction = ""
        concept_details_note = ""
        structure_note = ""
        prev_summary_note = ""
        instruction_lang_note = ""
        mermaid_lang_note = ""
        code_comment_note = ""
        link_lang_note = ""
        tone_note = ""
        if language.lower() != "english":
            lang_cap = language.capitalize()
            language_instruction = f"IMPORTANT: Write this ENTIRE tutorial chapter in **{lang_cap}**. Some input context (like concept name, description, chapter list, previous summary) might already be in {lang_cap}, but you MUST translate ALL other generated content including explanations, examples, technical terms, and potentially code comments into {lang_cap}. DO NOT use English anywhere except in code syntax, required proper nouns, or when specified. The entire output MUST be in {lang_cap}.\n\n"
            concept_details_note = f" (Note: Provided in {lang_cap})"
            structure_note = f" (Note: Chapter names might be in {lang_cap})"
            prev_summary_note = f" (Note: This summary might be in {lang_cap})"
            instruction_lang_note = f" (in {lang_cap})"
            mermaid_lang_note = f" (Use {lang_cap} for labels/text if appropriate)"
            code_comment_note = f" (Translate to {lang_cap} if possible, otherwise keep minimal English for clarity)"
            link_lang_note = (
                f" (Use the {lang_cap} chapter title from the structure above)"
            )
            tone_note = f" (appropriate for {lang_cap} readers)"

        prompt = f"""
{language_instruction}Write a very beginner-friendly tutorial chapter (in Markdown format) for the project `{project_name}` about the concept: "{abstraction_name}". This is Chapter {chapter_num}.

Concept Details{concept_details_note}:
- Name: {abstraction_name}
- Description:
{abstraction_description}

Complete Tutorial Structure{structure_note}:
{item["full_chapter_listing"]}

Context from previous chapters{prev_summary_note}:
{previous_chapters_summary if previous_chapters_summary else "This is the first chapter."}

Relevant Code Snippets (Code itself remains unchanged):
{file_context_str if file_context_str else "No specific code snippets provided for this abstraction."}

Instructions for the chapter (Generate content in {language.capitalize()} unless specified otherwise):
- Start with a clear heading (e.g., `# Chapter {chapter_num}: {abstraction_name}`). Use the provided concept name.

- If this is not the first chapter, begin with a brief transition from the previous chapter{instruction_lang_note}, referencing it with a proper Markdown link using its name{link_lang_note}.

- Begin with a high-level motivation explaining what problem this abstraction solves{instruction_lang_note}. Start with a central use case as a concrete example. The whole chapter should guide the reader to understand how to solve this use case. Make it very minimal and friendly to beginners.

- If the abstraction is complex, break it down into key concepts. Explain each concept one-by-one in a very beginner-friendly way{instruction_lang_note}.

- Explain how to use this abstraction to solve the use case{instruction_lang_note}. Give example inputs and outputs for code snippets (if the output isn't values, describe at a high level what will happen{instruction_lang_note}).

- Each code block should be BELOW 10 lines! If longer code blocks are needed, break them down into smaller pieces and walk through them one-by-one. Aggresively simplify the code to make it minimal. Use comments{code_comment_note} to skip non-important implementation details. Each code block should have a beginner friendly explanation right after it{instruction_lang_note}.

- Describe the internal implementation to help understand what's under the hood{instruction_lang_note}. First provide a non-code or code-light walkthrough on what happens step-by-step when the abstraction is called{instruction_lang_note}. It's recommended to use a simple sequenceDiagram with a dummy example - keep it minimal with at most 5 participants to ensure clarity. If participant name has space, use: `participant QP as Query Processing`. {mermaid_lang_note}.

- Then dive deeper into code for the internal implementation with references to files. Provide example code blocks, but make them similarly simple and beginner-friendly. Explain{instruction_lang_note}.

- IMPORTANT: When you need to refer to other core abstractions covered in other chapters, ALWAYS use proper Markdown links like this: [Chapter Title](filename.md). Use the Complete Tutorial Structure above to find the correct filename and the chapter title{link_lang_note}. Translate the surrounding text.

- Use mermaid diagrams to illustrate complex concepts (```mermaid``` format). {mermaid_lang_note}.

- Heavily use analogies and examples throughout{instruction_lang_note} to help beginners understand.

- End the chapter with a brief conclusion that summarizes what was learned{instruction_lang_note} and provides a transition to the next chapter{instruction_lang_note}. If there is a next chapter, use a proper Markdown link: [Next Chapter Title](next_chapter_filename){link_lang_note}.

- Ensure the tone is welcoming and easy for a newcomer to understand{tone_note}.

- Output *only* the Markdown content for this chapter.

Now, directly provide a super beginner-friendly Markdown output (DON'T need ```markdown``` tags):
"""
        chapter_content = call_llm(prompt, use_cache=(use_cache and self.cur_retry == 0)) # Use cache only if enabled and not retrying
        # Basic validation/cleanup
        actual_heading = f"# Chapter {chapter_num}: {abstraction_name}"  # Use potentially translated name
        if not chapter_content.strip().startswith(f"# Chapter {chapter_num}"):
            # Add heading if missing or incorrect, trying to preserve content
            lines = chapter_content.strip().split("\n")
            if lines and lines[0].strip().startswith(
                "#"
            ):  # If there's some heading, replace it
                lines[0] = actual_heading
                chapter_content = "\n".join(lines)
            else:  # Otherwise, prepend it
                chapter_content = f"{actual_heading}\n\n{chapter_content}"

        # Add the generated content to our temporary list for the next iteration's context
        self.chapters_written_so_far.append(chapter_content)

        return chapter_content  # Return the Markdown string (potentially translated)

    def post(self, shared, prep_res, exec_res_list):
        # exec_res_list contains the generated Markdown for each chapter, in order
        shared["chapters"] = exec_res_list
        # Clean up the temporary instance variable
        del self.chapters_written_so_far
        print(f"Finished writing {len(exec_res_list)} chapters.")


class CombineTutorial1(Node):
    def prep(self, shared):
        project_name = shared["project_name"]
        output_base_dir = shared.get("output_dir", "output")  # Default output dir
        output_path = os.path.join(output_base_dir, project_name)
        repo_url = shared.get("repo_url")  # Get the repository URL
        # language = shared.get("language", "english") # No longer needed for fixed strings

        # Get potentially translated data
        relationships_data = shared[
            "relationships"
        ]  # {"summary": str, "details": [{"from": int, "to": int, "label": str}]} -> summary/label potentially translated
        chapter_order = shared["chapter_order"]  # indices
        abstractions = shared[
            "abstractions"
        ]  # list of dicts -> name/description potentially translated
        chapters_content = shared[
            "chapters"
        ]  # list of strings -> content potentially translated

        # --- Generate Mermaid Diagram ---
        mermaid_lines = ["flowchart TD"]
        # Add nodes for each abstraction using potentially translated names
        for i, abstr in enumerate(abstractions):
            node_id = f"A{i}"
            # Use potentially translated name, sanitize for Mermaid ID and label
            sanitized_name = abstr["name"].replace('"', "")
            node_label = sanitized_name  # Using sanitized name only
            mermaid_lines.append(
                f'    {node_id}["{node_label}"]'
            )  # Node label uses potentially translated name
        # Add edges for relationships using potentially translated labels
        for rel in relationships_data["details"]:
            from_node_id = f"A{rel['from']}"
            to_node_id = f"A{rel['to']}"
            # Use potentially translated label, sanitize
            edge_label = (
                rel["label"].replace('"', "").replace("\n", " ")
            )  # Basic sanitization
            max_label_len = 30
            if len(edge_label) > max_label_len:
                edge_label = edge_label[: max_label_len - 3] + "..."
            mermaid_lines.append(
                f'    {from_node_id} -- "{edge_label}" --> {to_node_id}'
            )  # Edge label uses potentially translated label

        mermaid_diagram = "\n".join(mermaid_lines)
        # --- End Mermaid ---

        # --- Prepare index.md content ---
        index_content = f"# Tutorial: {project_name}\n\n"
        index_content += f"{relationships_data['summary']}\n\n"  # Use the potentially translated summary directly
        # Keep fixed strings in English
        index_content += f"**Source Repository:** [{repo_url}]({repo_url})\n\n"

        # Add Mermaid diagram for relationships (diagram itself uses potentially translated names/labels)
        index_content += "```mermaid\n"
        index_content += mermaid_diagram + "\n"
        index_content += "```\n\n"

        # Keep fixed strings in English
        index_content += f"## Chapters\n\n"

        chapter_files = []
        # Generate chapter links based on the determined order, using potentially translated names
        for i, abstraction_index in enumerate(chapter_order):
            # Ensure index is valid and we have content for it
            if 0 <= abstraction_index < len(abstractions) and i < len(chapters_content):
                abstraction_name = abstractions[abstraction_index][
                    "name"
                ]  # Potentially translated name
                # Sanitize potentially translated name for filename
                safe_name = "".join(
                    c if c.isalnum() else "_" for c in abstraction_name
                ).lower()
                filename = f"{i+1:02d}_{safe_name}.md"
                index_content += f"{i+1}. [{abstraction_name}]({filename})\n"  # Use potentially translated name in link text

                # Add attribution to chapter content (using English fixed string)
                chapter_content = chapters_content[i]  # Potentially translated content
                if not chapter_content.endswith("\n\n"):
                    chapter_content += "\n\n"
                # Keep fixed strings in English
                chapter_content += f"---\n\nGenerated by [AI Codebase Knowledge Builder](https://github.com/The-Pocket/Tutorial-Codebase-Knowledge)"

                # Store filename and corresponding content
                chapter_files.append({"filename": filename, "content": chapter_content})
            else:
                print(
                    f"Warning: Mismatch between chapter order, abstractions, or content at index {i} (abstraction index {abstraction_index}). Skipping file generation for this entry."
                )

        # Add attribution to index content (using English fixed string)
        index_content += f"\n\n---\n\nGenerated by [AI Codebase Knowledge Builder](https://github.com/The-Pocket/Tutorial-Codebase-Knowledge)"

        return {
            "output_path": output_path,
            "index_content": index_content,
            "chapter_files": chapter_files,  # List of {"filename": str, "content": str}
        }

    def exec(self, prep_res):
        output_path = prep_res["output_path"]
        index_content = prep_res["index_content"]
        chapter_files = prep_res["chapter_files"]

        print(f"Combining tutorial into directory: {output_path}")
        # Rely on Node's built-in retry/fallback
        os.makedirs(output_path, exist_ok=True)

        # Write index.md
        index_filepath = os.path.join(output_path, "index.md")
        with open(index_filepath, "w", encoding="utf-8") as f:
            f.write(index_content)
        print(f"  - Wrote {index_filepath}")

        # Write chapter files
        for chapter_info in chapter_files:
            chapter_filepath = os.path.join(output_path, chapter_info["filename"])
            with open(chapter_filepath, "w", encoding="utf-8") as f:
                f.write(chapter_info["content"])
            print(f"  - Wrote {chapter_filepath}")

        return output_path  # Return the final path

    def post(self, shared, prep_res, exec_res):
        shared["final_output_dir"] = exec_res  # Store the output path
        print(f"\nTutorial generation complete! Files are in: {exec_res}")


class AnalyzeCode(Node):
    def prep(self, shared):
        files_data = shared["files"]
        project_name = shared["project_name"]
        use_cache = shared.get("use_cache", True)
        
        # Get Excel validation data
        excel_validation = shared.get("excel_validation", {})
        unanswered_mandatory = excel_validation.get("unanswered_mandatory", [])
        
        # Get component questions from the Excel file
        component_questions = excel_validation.get("component_questions", {})
        
        def create_llm_context(files_data):
            context = ""
            file_info = []
            for i, (path, content) in enumerate(files_data):
                entry = f"--- File Index {i}: {path} ---\n{content}\n\n"
                context += entry
                file_info.append((i, path))
            return context, file_info

        context, file_info = create_llm_context(files_data)
        file_listing = "\n".join([f"- {idx} # {path}" for idx, path in file_info])
        
        # Create a mapping of file paths to their content for validation
        file_content_map = {path: content for path, content in files_data}
        
        return context, file_listing, len(files_data), project_name, use_cache, file_content_map, component_questions

    def exec(self, prep_res):
        context, file_listing, file_count, project_name, use_cache, file_content_map, component_questions = prep_res
        print(f"Analyzing code for comprehensive review...")

        # Create a list of components to check for in the codebase
        component_check_list = """
1. Venafi
2. Redis
3. Channel Secure / PingFed
4. NAS / SMB
5. SMTP
6. AutoSys
7. CRON/quartz/spring batch
8. MTLS / Mutual Auth / Hard Rock pattern
9. NDM
10. Legacy JKS files
11. SOAP Calls
12. REST API
13. APIGEE
14. KAFKA
15. IBM MQ
16. LDAP
17. Splunk
18. AppD / AppDynamics
19. ELASTIC APM
20. Harness or UCD for CI/CD
21. Hashicorp vault
22. Bridge Utility server
23. RabbitMQ
24. Databases:
   - MongoDB
   - SQLServer
   - MySQL
   - PostgreSQL
   - Oracle
   - Cassandra
   - Couchbase
   - Neo4j
   - Hadoop
   - Spark
25. Authentication methods:
   - Okta
   - SAML
   - Auth
   - JWT
   - OpenID
   - ADFS
"""

        # List of security and quality practices to check for
        security_checks = """
### Auditability
1. Avoid logging confidential data: Scan logs and code for PII/PHI or hardcoded secrets.
2. Create audit trail logs: Check for the presence of audit logs in the codebase.
3. Tracking ID for log messages: Verify if correlation/tracking IDs are used in log messages.
4. Log REST API calls: Inspect for middleware or interceptors that log API calls.
5. Log application messages: Look for usage of logger.info/warn/error patterns in the application.
6. Client UI errors are logged: Review frontend (JavaScript/TypeScript) for error logging handlers.

### Availability
1. Retry Logic: Identify retry patterns in HTTP clients or libraries.
2. Set timeouts on I/O operations: Look for timeout settings in HTTP, database, or file I/O operations.
3. Throttling, drop request: Detect rate limiter or logic that drops excessive requests.
4. Circuit breakers on outgoing requests: Identify use of circuit breaker libraries such as Hystrix or Resilience4j.

### Error Handling
1. Log system errors: Search for backend error logging patterns.
2. Use HTTP standard error codes: Check API responses for standard HTTP status codes.
3. Include client error tracking: Look for client-side error tracking using libraries like Sentry or custom implementations.

### Monitoring
1. URL monitoring: Detect health check or ping endpoints used for availability monitoring.

### Testing
1. Automated Regression Testing: Verify the presence of automated test suites in the project.
"""

        # Add excel component questions to prompt if available
        excel_component_section = ""
        if component_questions:
            excel_component_section = """
IMPORTANT: The following components were mentioned in the project intake form. 
Please pay special attention to detecting these components in the codebase:

"""
            for component, data in component_questions.items():
                answer = "Yes" if data.get("is_yes", False) else "No"
                excel_component_section += f"- {component}: Declared as '{answer}' in the intake form\n"

        prompt = f"""
You are a code review assistant trained to identify patterns in source code and configurations related to various aspects of code quality, security, and best practices.

Analyze the following codebase for the project `{project_name}`:

Codebase Context:
{context}

List of files:
{file_listing}

{excel_component_section}

IMPORTANT: You MUST return a JSON object with FOUR main sections:
1. "technology_stack": A dictionary containing all technologies found in the codebase
2. "findings": A list of issues and recommendations
3. "component_analysis": A dictionary containing detected components with yes/no values
4. "security_quality_analysis": A dictionary containing results of security and quality practice checks

For the technology stack, you MUST identify and categorize ALL technologies used in the codebase. Look for:
- Programming languages (e.g., Python, JavaScript, Java)
- Frameworks (e.g., React, Django, Spring)
- Libraries (e.g., pandas, numpy, lodash)
- Databases (e.g., PostgreSQL, MongoDB)
- Tools (e.g., Docker, Kubernetes)
- Services (e.g., AWS, Azure)

For each technology, provide:
- name: The technology name
- version: The version if found, or "unknown"
- purpose: What it's used for in this codebase
- files: List of files where it's used - YOU MUST INCLUDE ACTUAL FILE PATHS

Example technology stack format:
```json
{{
    "technology_stack": {{
        "programming_languages": [
            {{
                "name": "Python",
                "version": "3.8",
                "purpose": "Main application language",
                "files": ["main.py", "utils.py", "core/genflow.py"]
            }}
        ],
        "frameworks": [
            {{
                "name": "Flask",
                "version": "2.0",
                "purpose": "Web framework",
                "files": ["app.py", "routes/api.py"]
            }}
        ]
    }}
}}
```

Additionally, specifically analyze the codebase to check for the presence of the following components. For each, respond with "yes" if found, or "no" if not found, with evidence from the code if available:

{component_check_list}

For the component_analysis section, use this format:
```json
{{
    "component_analysis": {{
        "venafi": {{
            "detected": "yes",
            "evidence": "Found Venafi certificate management in security/certs.py (line 42)"
        }},
        "redis": {{
            "detected": "no",
            "evidence": "No Redis dependencies or configurations found"
        }}
        // ... and so on for all components
    }}
}}
```

Next, specifically analyze the codebase for the following security and quality practices:

{security_checks}

For the security_quality_analysis section, use this format with the actual data:
```json
{{
    "security_quality_analysis": {{
        "auditability": {{
            "avoid_logging_confidential_data": {{
                "implemented": "yes",
                "evidence": "Found proper masking of sensitive data in logging/utils.py (line 25)",
                "recommendation": "None needed; good practices observed."
            }},
            "create_audit_trail_logs": {{
                "implemented": "no",
                "evidence": "No evidence of comprehensive audit trails found",
                "recommendation": "Implement centralized audit logging for all critical actions"
            }}
            // ... and so on for all security practices
        }},
        "availability": {{
            // Details for availability checks
        }},
        "error_handling": {{
            // Details for error handling checks
        }},
        "monitoring": {{
            // Details for monitoring checks
        }},
        "testing": {{
            // Details for testing checks
        }}
    }}
}}
```

For each security and quality practice check, include:
- implemented: "yes", "partial", or "no"
- evidence: Detailed evidence from the code, including specific file names and line numbers where applicable
- recommendation: What should be done to improve or implement the practice

Then proceed with the existing best practice checks and return findings in the same format as before.

IMPORTANT RULES:
1. You MUST identify and return the technology stack, even if you find no issues
2. For each technology, you MUST provide at least one file where it's used with SPECIFIC FILE PATHS
3. All evidence should include SPECIFIC FILE PATHS and line numbers when possible
4. If you can't determine a version, use "unknown"
5. If you can't determine a purpose, provide a general description
6. The response MUST be valid JSON that can be parsed by json.loads()
7. Do not include any text before or after the JSON object
8. Do not use comments in the JSON response
9. Use double quotes for all strings in the JSON response
10. You MUST include ALL the security and quality practice checks in the response, even if not implemented

Now, analyze the codebase and return the complete JSON response:
"""

        response = call_llm(prompt, use_cache=(use_cache and self.cur_retry == 0))
        
        # Extract and parse JSON from response
        try:
            # Clean the response - remove any non-JSON content
            response = response.strip()
            
            # Find the first { and last }
            start_idx = response.find('{')
            end_idx = response.rfind('}') + 1
            
            if start_idx == -1 or end_idx == 0:
                print("Warning: No JSON object found in response. Returning empty findings.")
                print("Response was:", response)
                return {"technology_stack": {}, "findings": [], "component_analysis": {}, "security_quality_analysis": {}}
                
            json_str = response[start_idx:end_idx]
            
            # Remove any comments from the JSON string
            json_str = re.sub(r'//.*?\n', '\n', json_str)
            
            # Try to parse the JSON
            try:
                analysis = json.loads(json_str)
            except json.JSONDecodeError as e:
                print(f"Warning: Failed to parse JSON response: {str(e)}")
                print("JSON string was:", json_str)
                return {"technology_stack": {}, "findings": [], "component_analysis": {}, "security_quality_analysis": {}}
            
            # Validate analysis structure
            if not isinstance(analysis, dict):
                print("Warning: Response is not a dictionary. Returning empty findings.")
                print("Response was:", analysis)
                return {"technology_stack": {}, "findings": [], "component_analysis": {}, "security_quality_analysis": {}}
                
            # Ensure technology_stack exists and is a dict
            if "technology_stack" not in analysis or not isinstance(analysis["technology_stack"], dict):
                analysis["technology_stack"] = {}
                
            # Ensure findings exists and is a list
            if "findings" not in analysis or not isinstance(analysis["findings"], list):
                analysis["findings"] = []
                
            # Ensure component_analysis exists and is a dict
            if "component_analysis" not in analysis or not isinstance(analysis["component_analysis"], dict):
                analysis["component_analysis"] = {}
                
            # Ensure security_quality_analysis exists and is a dict
            if "security_quality_analysis" not in analysis or not isinstance(analysis["security_quality_analysis"], dict):
                analysis["security_quality_analysis"] = {}
            
            # Add Excel component questions to analysis result
            if component_questions:
                analysis["excel_components"] = component_questions
            
            # Validate and clean up technology stack
            validated_tech_stack = {}
            for category, techs in analysis["technology_stack"].items():
                if not isinstance(techs, list):
                    continue
                    
                validated_techs = []
                for tech in techs:
                    if not isinstance(tech, dict):
                        continue
                        
                    # Ensure required fields exist
                    if "name" not in tech:
                        continue
                        
                    # Add missing fields with defaults
                    tech = {
                        "name": tech["name"],
                        "version": tech.get("version", "unknown"),
                        "purpose": tech.get("purpose", "N/A"),
                        "files": tech.get("files", [])
                    }
                    
                    # Validate files exist
                    valid_files = []
                    for file_path in tech["files"]:
                        if file_path in file_content_map:
                            valid_files.append(file_path)
                    
                    tech["files"] = valid_files
                    validated_techs.append(tech)
                
                if validated_techs:
                    validated_tech_stack[category] = validated_techs
            
            # Validate findings
            validated_findings = []
            for finding in analysis["findings"]:
                if not isinstance(finding, dict):
                    continue
                    
                # Check required fields
                if not all(k in finding for k in ["category", "severity", "description", "location", "recommendation"]):
                    continue
                    
                # Validate location
                location = finding["location"]
                if not isinstance(location, dict) or not all(k in location for k in ["file", "line", "code"]):
                    continue
                    
                # Validate types
                if not isinstance(finding["category"], str) or \
                   not isinstance(finding["severity"], str) or \
                   not isinstance(finding["description"], str) or \
                   not isinstance(finding["recommendation"], str) or \
                   not isinstance(location["file"], str) or \
                   not isinstance(location["code"], str):
                    continue
                    
                # Convert line to int if it's a string
                try:
                    location["line"] = int(location["line"])
                except (ValueError, TypeError):
                    continue
                
                # Validate file exists and line number is valid
                file_path = location["file"]
                if file_path not in file_content_map:
                    continue
                    
                file_content = file_content_map[file_path]
                file_lines = file_content.split('\n')
                
                if not (0 <= location["line"] < len(file_lines)):
                    continue
                
                # Validate code snippet matches the actual line
                actual_line = file_lines[location["line"]].strip()
                reported_snippet = location["code"].strip()
                
                if reported_snippet not in actual_line:
                    continue
                        
                validated_findings.append(finding)
            
            # Validate component analysis
            validated_component_analysis = {}
            for component, analysis_data in analysis.get("component_analysis", {}).items():
                if not isinstance(analysis_data, dict):
                    continue
                
                if "detected" not in analysis_data:
                    continue
                
                detected = analysis_data["detected"].lower()
                if detected not in ["yes", "no"]:
                    detected = "no"
                
                evidence = analysis_data.get("evidence", "No evidence provided")
                
                validated_component_analysis[component] = {
                    "detected": detected,
                    "evidence": evidence
                }
            
            # Validate security and quality analysis
            validated_security_quality = {}
            
            # Define expected categories and practices
            expected_security_structure = {
                "auditability": [
                    "avoid_logging_confidential_data",
                    "create_audit_trail_logs",
                    "tracking_id_for_log_messages",
                    "log_rest_api_calls",
                    "log_application_messages",
                    "client_ui_errors_are_logged"
                ],
                "availability": [
                    "retry_logic",
                    "set_timeouts_on_io_operations",
                    "throttling_drop_request",
                    "circuit_breakers_on_outgoing_requests"
                ],
                "error_handling": [
                    "log_system_errors",
                    "use_http_standard_error_codes",
                    "include_client_error_tracking"
                ],
                "monitoring": [
                    "url_monitoring"
                ],
                "testing": [
                    "automated_regression_testing"
                ]
            }
            
            security_quality = analysis.get("security_quality_analysis", {})
            
            # For each expected category, validate the practices
            for category, practices in expected_security_structure.items():
                category_data = security_quality.get(category, {})
                validated_practices = {}
                
                for practice in practices:
                    practice_data = category_data.get(practice, {})
                    
                    if not isinstance(practice_data, dict):
                        practice_data = {
                            "implemented": "no",
                            "evidence": "Not analyzed",
                            "recommendation": "Implement this practice"
                        }
                    
                    # Ensure all required fields are present
                    implemented = practice_data.get("implemented", "no").lower()
                    if implemented not in ["yes", "no", "partial"]:
                        implemented = "no"
                        
                    evidence = practice_data.get("evidence", "No evidence provided")
                    recommendation = practice_data.get("recommendation", "Implement this practice")
                    
                    validated_practices[practice] = {
                        "implemented": implemented,
                        "evidence": evidence,
                        "recommendation": recommendation
                    }
                
                validated_security_quality[category] = validated_practices
            
            print(f"Found {sum(len(techs) for techs in validated_tech_stack.values())} technologies, {len(validated_findings)} findings, and {len(validated_component_analysis)} components.")
            print(f"Analyzed {len(validated_security_quality)} security and quality categories.")
            
            # Compare component analysis from code with Excel answers
            if component_questions:
                print("\nComparing components found in code with Excel declarations:")
                for component_name, data in validated_component_analysis.items():
                    for excel_comp, excel_data in component_questions.items():
                        if component_name.lower() in excel_comp.lower() or excel_comp.lower() in component_name.lower():
                            excel_answer = "Yes" if excel_data.get("is_yes", False) else "No"
                            code_answer = "Yes" if data["detected"].lower() == "yes" else "No"
                            match_str = "MATCH" if excel_answer == code_answer else "MISMATCH"
                            print(f"- {component_name}: Excel={excel_answer}, Code={code_answer} => {match_str}")
            
            return {
                "technology_stack": validated_tech_stack,
                "findings": validated_findings,
                "component_analysis": validated_component_analysis,
                "excel_components": component_questions if component_questions else {},
                "security_quality_analysis": validated_security_quality
            }
            
        except Exception as e:
            print(f"Warning: Unexpected error processing response: {str(e)}")
            print("Response was:", response)
            return {"technology_stack": {}, "findings": [], "component_analysis": {}, "security_quality_analysis": {}}

    def post(self, shared, prep_res, exec_res):
        print("\nDEBUG: AnalyzeCode post")
        print("DEBUG: Analysis results:", exec_res)
        print("DEBUG: Technology stack found:", list(exec_res.get("technology_stack", {}).keys()))
        print("DEBUG: Number of findings:", len(exec_res.get("findings", [])))
        print("DEBUG: Components detected:", len(exec_res.get("component_analysis", {})))
        print("DEBUG: Security and quality checks:", list(exec_res.get("security_quality_analysis", {}).keys()))
        shared["code_analysis"] = exec_res
        print("DEBUG: Stored code_analysis in shared state")


class GenerateReport(Node):
    def prep(self, shared):
        print("\nDEBUG: Starting GenerateReport prep")
        print("DEBUG: Keys in shared:", list(shared.keys()))
        
        analysis = shared.get("code_analysis", {})
        print("DEBUG: Code analysis data:", analysis)
        
        findings = analysis.get("findings", [])
        print("DEBUG: Number of findings:", len(findings))
        
        # Support both old and new technology formats
        if "technology_stack" in analysis:
            technology_stack = analysis["technology_stack"]
        else:
            technology_stack = analysis.get("technologies", {})
        print("DEBUG: Technology stack:", technology_stack)
        
        # Get component analysis
        component_analysis = analysis.get("component_analysis", {})
        print("DEBUG: Component analysis:", component_analysis)
        
        # Get Excel component declarations
        excel_components = analysis.get("excel_components", {})
        print("DEBUG: Excel component declarations:", excel_components)
        
        # Get security and quality analysis
        security_quality_analysis = analysis.get("security_quality_analysis", {})
        print("DEBUG: Security and quality analysis:", list(security_quality_analysis.keys()))
        
        # Get Jira stories if available
        jira_stories = shared.get("jira_stories", [])
        print("DEBUG: Number of Jira stories:", len(jira_stories))
        
        project_name = shared.get("project_name", "Unknown Project")
        output_dir = shared.get("output_dir", "analysis_output")
        
        # Get Excel validation data if available
        excel_validation = shared.get("excel_validation", {})
        
        print("DEBUG: Project name:", project_name)
        print("DEBUG: Output directory:", output_dir)
        
        return findings, technology_stack, project_name, output_dir, excel_validation, component_analysis, excel_components, security_quality_analysis, jira_stories

    def exec(self, prep_res):
        findings, technology_stack, project_name, output_dir, excel_validation, component_analysis, excel_components, security_quality_analysis, jira_stories = prep_res
        print("\nDEBUG: Starting GenerateReport exec")
        print("DEBUG: Number of findings:", len(findings))
        print("DEBUG: Technology stack categories:", list(technology_stack.keys()))
        print("DEBUG: Component analysis items:", len(component_analysis))
        print("DEBUG: Security analysis categories:", list(security_quality_analysis.keys()))
        
        # Initialize report and html_content variables at the beginning of the method
        # Generate Markdown report content
        report = f"# Code Analysis Report for {project_name}\n\n"
        
        # Add Executive Summary with statistics
        report += "## Executive Summary\n\n"
        
        # Calculate statistics for executive summary
        security_stats = ""
        if security_quality_analysis:
            categories_total = 0
            categories_implemented = 0
            categories_partial = 0
            categories_not_implemented = 0
            
            for category, practices in security_quality_analysis.items():
                for practice, details in practices.items():
                    categories_total += 1
                    status = details.get("implemented", "no")
                    if status == "yes":
                        categories_implemented += 1
                    elif status == "partial":
                        categories_partial += 1
                    else:
                        categories_not_implemented += 1
            
            if categories_total > 0:
                implementation_percentage = ((categories_implemented + 0.5 * categories_partial) / categories_total) * 100
                security_stats = f"- **Security & Quality Implementation**: {implementation_percentage:.1f}%\n"
                security_stats += f"- **Practices Implemented**: {categories_implemented} fully, {categories_partial} partially, {categories_not_implemented} not implemented\n"
        
        # Calculate findings statistics
        findings_stats = ""
        if findings:
            severity_counts = {"High": 0, "Medium": 0, "Low": 0}
            for finding in findings:
                severity = finding.get("severity", "Low")
                if severity in severity_counts:
                    severity_counts[severity] += 1
            
            findings_stats = f"- **Total Findings**: {len(findings)}\n"
            if severity_counts["High"] > 0:
                findings_stats += f"- **High Severity Issues**: {severity_counts['High']}\n"
            if severity_counts["Medium"] > 0:
                findings_stats += f"- **Medium Severity Issues**: {severity_counts['Medium']}\n"
            if severity_counts["Low"] > 0:
                findings_stats += f"- **Low Severity Issues**: {severity_counts['Low']}\n"
        else:
            findings_stats = "- **Total Findings**: 0\n"
        
        # Calculate component mismatches
        component_stats = ""
        mismatches_count = 0
        if excel_components and component_analysis:
            for excel_comp, excel_data in excel_components.items():
                excel_comp_lower = excel_comp.lower()
                excel_declared = excel_data.get("is_yes", False)
                
                for comp_name, comp_data in component_analysis.items():
                    if excel_comp_lower in comp_name.lower() or comp_name.lower() in excel_comp_lower:
                        detected = comp_data["detected"].lower() == "yes"
                        if excel_declared != detected:
                            mismatches_count += 1
                        break
                        
            if mismatches_count > 0:
                component_stats = f"- **Component Declaration Mismatches**: {mismatches_count}\n"
        
        # Add all statistics to executive summary
        report += security_stats + findings_stats + component_stats + "\n"
        
        # Add Intake Form Validation Section if available
        if excel_validation:
            report += "## Intake Form Validation\n\n"
            
            # Check if the form is valid
            is_valid = excel_validation.get("is_valid", False)
            total_rows = excel_validation.get("total_rows", 0)
            mandatory_fields = excel_validation.get("mandatory_fields", 0)
            unanswered_mandatory = excel_validation.get("unanswered_mandatory", [])
            git_repo_url = excel_validation.get("git_repo_url", "Not provided")
            git_repo_valid = excel_validation.get("git_repo_valid", False)
            
            if is_valid:
                report += "✅ **Intake form is complete.** All mandatory fields have been answered.\n\n"
            else:
                report += "❌ **Intake form is incomplete.** Some mandatory fields have not been answered or Git repository is invalid.\n\n"
                
            report += f"- Total questions: {total_rows}\n"
            report += f"- Mandatory fields: {mandatory_fields}\n"
            report += f"- Unanswered mandatory fields: {len(unanswered_mandatory)}\n"
            report += f"- Git repository URL: {git_repo_url}\n"
            report += f"- Git repository valid: {'✅ Yes' if git_repo_valid else '❌ No'}\n\n"
            
            if unanswered_mandatory:
                report += "### Unanswered Mandatory Questions\n\n"
                for question in unanswered_mandatory:
                    report += f"- {question}\n"
                report += "\n"
        
        # Add Table of Contents
        report += "## Table of Contents\n\n"
        report += "1. [Executive Summary](#executive-summary)\n"
        report += "2. [Component Analysis](#component-analysis)\n"
        report += "3. [Security and Quality Practices](#security-and-quality-practices)\n"
        report += "4. [Technology Stack](#technology-stack)\n"
        
        if jira_stories:
            report += "5. [Jira Stories](#jira-stories)\n"
            report += "6. [Action Items](#action-items)\n"
        else:
            report += "5. [Action Items](#action-items)\n"
        
        # Component Analysis Section
        report += "## Component Analysis\n\n"
        
        if excel_components and component_analysis:
            report += "The following table compares the components declared in the intake form with the components detected in the codebase:\n\n"
            
            # Table header
            report += "| Component | Declared | Detected | Status |\n"
            report += "|-----------|----------|----------|--------|\n"
            
            for excel_comp, excel_data in excel_components.items():
                excel_comp_lower = excel_comp.lower()
                excel_declared = excel_data.get("is_yes", False)
                
                # Find if component is in the analysis
                matched = False
                detected = False
                
                for comp_name, comp_data in component_analysis.items():
                    if excel_comp_lower in comp_name.lower() or comp_name.lower() in excel_comp_lower:
                        matched = True
                        detected = comp_data["detected"].lower() == "yes"
                        break
                
                status = "Match" if excel_declared == detected else "Mismatch"
                
                report += f"| {excel_comp} | {'Yes' if excel_declared else 'No'} | {'Yes' if detected else 'No'} | {status} |\n"
                
            report += "\n"
        elif component_analysis:
            report += "The following components were detected in the codebase:\n\n"
            
            # Table header
            report += "| Component | Detected | Evidence |\n"
            report += "|-----------|----------|----------|\n"
            
            for comp_name, comp_data in component_analysis.items():
                detected = comp_data.get("detected", "no")
                evidence = comp_data.get("evidence", "None provided")
                
                # Truncate evidence if too long
                if len(evidence) > 50:
                    evidence = evidence[:47] + "..."
                
                report += f"| {comp_name} | {detected} | {evidence} |\n"
                
            report += "\n"
        else:
            report += "No component analysis data available.\n\n"
        
        # Security and Quality Practices Section
        report += "## Security and Quality Practices\n\n"
        
        if security_quality_analysis:
            report += "The following sections summarize the security and quality practices implemented in the codebase:\n\n"
            
            for category, practices in security_quality_analysis.items():
                # Convert category from snake_case to Title Case
                category_display = category.replace('_', ' ').title()
                report += f"### {category_display}\n\n"
                
                # Table header
                report += "| Practice | Status | Evidence |\n"
                report += "|----------|--------|----------|\n"
                
                for practice_name, practice_data in practices.items():
                    # Convert practice name from snake_case to Title Case
                    practice_display = practice_name.replace('_', ' ').title()
                    
                    status = practice_data.get("implemented", "no")
                    evidence = practice_data.get("evidence", "None provided")
                    
                    status_text = "Implemented" if status == "yes" else "Partially Implemented" if status == "partial" else "Not Implemented"
                    
                    # Truncate evidence if too long
                    if len(evidence) > 50:
                        evidence = evidence[:47] + "..."
                    
                    report += f"| {practice_display} | {status_text} | {evidence} |\n"
                
                report += "\n"
        else:
            report += "No security and quality practice analysis data available.\n\n"
        
        # Technology Stack Section
        report += "## Technology Stack\n\n"
        
        if technology_stack:
            report += "The following technologies were identified in the codebase:\n\n"
            
            # Group technologies by category
            for category, techs in technology_stack.items():
                if not isinstance(techs, list) or not techs:
                    continue
                    
                # Convert category from snake_case to Title Case
                category_display = category.replace('_', ' ').title()
                report += f"### {category_display}\n\n"
                
                # Create table for this category
                report += "| Technology | Version | Purpose | Files |\n"
                report += "|------------|---------|---------|-------|\n"
                
                for tech in techs:
                    if not isinstance(tech, dict):
                        continue
                        
                    tech_name = tech.get("name", "Unknown")
                    tech_version = tech.get("version", "unknown")
                    tech_purpose = tech.get("purpose", "Not specified")
                    tech_files = tech.get("files", [])
                    
                    # Truncate purpose if too long (for table formatting)
                    if len(tech_purpose) > 25:
                        tech_purpose = tech_purpose[:22] + "..."
                        
                    # Format files list
                    files_display = ", ".join(tech_files[:2])
                    if len(tech_files) > 2:
                        files_display += f" (+{len(tech_files) - 2})"
                    
                    report += f"| {tech_name} | {tech_version} | {tech_purpose} | {files_display} |\n"
                
                report += "\n"
        else:
            report += "No technologies were identified in the codebase.\n\n"
        
        # Detailed Findings Section
        if findings:
            report += "## Detailed Findings\n\n"
            
            # Group findings by category
            categories = {}
            for finding in findings:
                category = finding.get("category", "Uncategorized")
                if category not in categories:
                    categories[category] = []
                categories[category].append(finding)
            
            for category, category_findings in categories.items():
                report += f"### {category}\n\n"
                
                for finding in category_findings:
                    severity = finding.get("severity", "Low")
                    description = finding.get("description", "No description provided")
                    recommendation = finding.get("recommendation", "No recommendation provided")
                    location = finding.get("location", {})
                    file_path = location.get("file", "Unknown")
                    line = location.get("line", "Unknown")
                    code = location.get("code", "No code snippet provided")
                    
                    report += f"#### {description}\n\n"
                    report += f"**Severity**: {severity}\n\n"
                    report += f"**Location**: {file_path}, line {line}\n\n"
                    report += f"**Code**: `{code}`\n\n"
                    report += f"**Recommendation**: {recommendation}\n\n"
        
        # Action Items Section
        report += "## Action Items\n\n"
        
        action_items = []
        
        # 1. Intake form completion if needed
        if excel_validation and not excel_validation.get("is_valid", True):
            action_items.append({
                "priority": "Critical",
                "title": "Complete Intake Form",
                "description": "Complete all mandatory questions in the intake form to ensure accurate project configuration."
            })
            
        # 2. High-priority findings
        high_findings = [f for f in findings if f.get("severity", "Low") == "High"]
        if high_findings:
            action_items.append({
                "priority": "High",
                "title": f"Address {len(high_findings)} High-Priority Issues",
                "description": "Fix critical security and quality issues to prevent potential vulnerabilities."
            })
            
        # 3. Component mismatches
        if mismatches_count > 0:
            action_items.append({
                "priority": "Medium",
                "title": "Resolve Component Discrepancies",
                "description": "Reconcile mismatches between declared components and what's detected in the code."
            })
            
        # 4. Missing security practices
        if security_quality_analysis:
            missing_practices_count = 0
            for category, practices in security_quality_analysis.items():
                for practice, details in practices.items():
                    if details.get("implemented", "no") == "no":
                        missing_practices_count += 1
                        
            if missing_practices_count > 0:
                action_items.append({
                    "priority": "Medium",
                    "title": f"Implement {missing_practices_count} Missing Security Practices",
                    "description": "Address security gaps to improve overall application security posture."
                })
                
        # 5. Medium-priority findings
        medium_findings = [f for f in findings if f.get("severity", "Low") == "Medium"]
        if medium_findings:
            action_items.append({
                "priority": "Medium",
                "title": f"Address {len(medium_findings)} Medium-Priority Issues",
                "description": "Resolve quality and security issues of moderate severity."
            })
        
        # Format action items in the report
        if action_items:
            for i, item in enumerate(action_items):
                report += f"### {i+1}. {item['title']} (Priority: {item['priority']})\n\n"
                report += f"{item['description']}\n\n"
        else:
            report += "No specific action items identified. The codebase appears to follow good practices.\n\n"

        # Add Jira Stories Section if available
        if jira_stories:
            report += "## Jira Stories\n\n"
            report += "The following Jira stories are relevant to this project:\n\n"
            
            for story in jira_stories:
                report += f"### {story['key']}: {story['summary']}\n\n"
                report += f"**Status**: {story['status']}\n\n"
                report += f"**Created**: {story['created']}\n\n"
                report += f"**Last Updated**: {story['updated']}\n\n"
                
                if story['description']:
                    report += f"**Description**:\n\n{story['description']}\n\n"
                
                if story['comments']:
                    report += "**Comments**:\n\n"
                    for comment in story['comments']:
                        report += f"- **{comment['author']}** ({comment['created']}):\n  {comment['body']}\n\n"
                
                if story['attachments']:
                    report += "**Attachments**:\n\n"
                    for attachment in story['attachments']:
                        report += f"- {attachment['filename']} ({attachment['size']} bytes)\n"
                    report += "\n"
            
            report += "\n"

        # Initialize HTML content outside the PDF generation try block
        # Create HTML content for the report
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Code Analysis Report - """ + project_name + """</title>
            <style>
                body {
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                    line-height: 1.4;
                    margin: 0;
                    padding: 15px;
                    color: #333;
                    max-width: 1200px;
                    margin: 0 auto;
                }
                h1 { font-size: 1.5em; margin: 0 0 15px 0; padding-bottom: 5px; border-bottom: 2px solid #eee; }
                h2 { font-size: 1.2em; margin: 15px 0 10px 0; color: #444; padding-top: 10px; border-top: 1px solid #eee; }
                h2:first-of-type { border-top: none; }
                h3 { font-size: 1.1em; margin: 10px 0 5px 0; color: #555; }
                .section {
                    margin: 10px 0;
                    padding: 10px;
                    background: #f8f9fa;
                    border-radius: 4px;
                }
                .component-table {
                    width: 100%;
                    border-collapse: collapse;
                    margin: 15px 0;
                }
                .component-table th, .component-table td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                .component-table th {
                    background-color: #f2f2f2;
                    font-weight: bold;
                }
                .component-table tr:nth-child(even) {
                    background-color: #f9f9f9;
                }
                .finding-list {
                    display: grid;
                    grid-template-columns: repeat(auto-fill, minmax(400px, 1fr));
                    gap: 10px;
                    margin: 10px 0;
                }
                .finding {
                    background: white;
                    padding: 10px;
                    border-radius: 4px;
                    border: 1px solid #eee;
                }
                .action-item {
                    background: #f8f9fa;
                    padding: 10px 15px;
                    margin: 10px 0;
                    border-radius: 4px;
                    border-left: 5px solid #007bff;
                }
                .action-item.critical {
                    border-left-color: #dc3545;
                }
                .action-item.high {
                    border-left-color: #fd7e14;
                }
                .action-item.medium {
                    border-left-color: #ffc107;
                }
                .action-item.low {
                    border-left-color: #28a745;
                }
                .severity-high { color: #dc3545; }
                .severity-medium { color: #fd7e14; }
                .severity-low { color: #28a745; }
                .severity-badge {
                    display: inline-block;
                    padding: 2px 6px;
                    border-radius: 3px;
                    font-size: 0.9em;
                    font-weight: 500;
                    margin-right: 8px;
                }
                .severity-high .severity-badge { background: #dc3545; color: white; }
                .severity-medium .severity-badge { background: #fd7e14; color: white; }
                .severity-low .severity-badge { background: #28a745; color: white; }
                .validation-status {
                    display: inline-block;
                    padding: 5px 10px;
                    border-radius: 4px;
                    font-weight: bold;
                }
                .validation-complete { background: #d4edda; color: #155724; }
                .validation-incomplete { background: #f8d7da; color: #721c24; }
                .component-yes {
                    color: #28a745;
                    font-weight: bold;
                }
                .component-no {
                    color: #dc3545;
                }
                .match {
                    color: #28a745;
                    font-weight: bold;
                }
                .mismatch {
                    color: #fd7e14;
                    font-weight: bold;
                }
                .toc {
                    background: #f8f9fa;
                    padding: 10px 15px;
                    border-radius: 4px;
                    margin: 15px 0;
                }
                .toc ul {
                    margin: 5px 0;
                    padding-left: 20px;
                }
                .executive-summary {
                    background: #e9f7fd;
                    border-radius: 4px;
                    padding: 10px 15px;
                    margin: 15px 0;
                    border-left: 5px solid #17a2b8;
                }
                pre, code {
                    background: #f8f9fa;
                    padding: 2px 4px;
                    border-radius: 3px;
                    font-family: monospace;
                    font-size: 0.9em;
                    overflow-x: auto;
                }
                .summary {
                    display: grid;
                    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                    gap: 10px;
                    margin: 10px 0;
                }
                .summary-item {
                    background: white;
                    padding: 10px;
                    border-radius: 4px;
                    border: 1px solid #eee;
                }
                .attachments {
                    display: grid;
                    grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
                    gap: 15px;
                    margin: 15px 0;
                }
                .attachment {
                    background: white;
                    padding: 10px;
                    border-radius: 4px;
                    border: 1px solid #eee;
                    text-align: center;
                }
                .attachment img {
                    margin: 10px auto;
                    display: block;
                    max-width: 100%;
                    height: auto;
                    border-radius: 3px;
                    box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                }
                ul, ol { margin: 5px 0; padding-left: 20px; }
                li { margin: 3px 0; }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin: 15px 0;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                th {
                    background-color: #f2f2f2;
                    font-weight: bold;
                }
                tr:nth-child(even) {
                    background-color: #f9f9f9;
                }
                @media print {
                    body { padding: 10px; }
                    .section { break-inside: avoid; }
                }
            </style>
        </head>
        <body>
            <h1>Code Analysis Report for """ + project_name + """</h1>
            
            <!-- Executive Summary -->
            <div class="executive-summary">
                <h2 id="executive-summary">Executive Summary</h2>
        """

        # Add security implementation stats to executive summary
        if security_quality_analysis:
            categories_total = 0
            categories_implemented = 0
            categories_partial = 0
            categories_not_implemented = 0
            
            for category, practices in security_quality_analysis.items():
                for practice, details in practices.items():
                    categories_total += 1
                    status = details.get("implemented", "no")
                    if status == "yes":
                        categories_implemented += 1
                    elif status == "partial":
                        categories_partial += 1
                    else:
                        categories_not_implemented += 1
            
            if categories_total > 0:
                implementation_percentage = ((categories_implemented + 0.5 * categories_partial) / categories_total) * 100
                html_content += f"""
                <ul>
                    <li><strong>Security & Quality Implementation</strong>: {implementation_percentage:.1f}%</li>
                    <li><strong>Practices Implemented</strong>: {categories_implemented} fully, {categories_partial} partially, {categories_not_implemented} not implemented</li>
                """
        
        # Add findings stats
        severity_counts = {"High": 0, "Medium": 0, "Low": 0}
        if findings:
            for finding in findings:
                severity = finding.get("severity", "Low")
                if severity in severity_counts:
                    severity_counts[severity] += 1
            
            html_content += f"""
                    <li><strong>Total Findings</strong>: {len(findings)}</li>
            """
            
            if severity_counts["High"] > 0:
                html_content += f"""
                    <li><strong>High Severity Issues</strong>: {severity_counts['High']}</li>
                """
            if severity_counts["Medium"] > 0:
                html_content += f"""
                    <li><strong>Medium Severity Issues</strong>: {severity_counts['Medium']}</li>
                """
            if severity_counts["Low"] > 0:
                html_content += f"""
                    <li><strong>Low Severity Issues</strong>: {severity_counts['Low']}</li>
                """
        else:
            html_content += f"""
                    <li><strong>Total Findings</strong>: 0</li>
            """
        
        # Add component mismatches
        mismatches_count = 0
        if excel_components and component_analysis:
            for excel_comp, excel_data in excel_components.items():
                excel_comp_lower = excel_comp.lower()
                excel_declared = excel_data.get("is_yes", False)
                
                for comp_name, comp_data in component_analysis.items():
                    if excel_comp_lower in comp_name.lower() or comp_name.lower() in excel_comp_lower:
                        detected = comp_data["detected"].lower() == "yes"
                        if excel_declared != detected:
                            mismatches_count += 1
                        break
                        
            if mismatches_count > 0:
                html_content += f"""
                    <li><strong>Component Declaration Mismatches</strong>: {mismatches_count}</li>
                """
        
        html_content += """
                </ul>
            </div>
        """
        
        # Add Table of Contents
        html_content += """
            <div class="toc">
                <h2 id="table-of-contents">Table of Contents</h2>
                <ol>
                    <li><a href="#executive-summary">Executive Summary</a></li>
        """
        
        toc_index = 2
        if excel_validation:
            html_content += """
                    <li><a href="#intake-form-validation">Intake Form Validation</a></li>
            """
            toc_index += 1
            
        html_content += """
                    <li><a href="#component-analysis">Component Analysis</a></li>
                    <li><a href="#security-and-quality-practices">Security and Quality Practices</a></li>
                    <li><a href="#technology-stack">Technology Stack</a></li>
        """
        
        if findings:
            html_content += """
                    <li><a href="#detailed-findings">Detailed Findings</a></li>
            """
        
        if jira_stories:
            html_content += """
                    <li><a href="#jira-stories">Jira Stories</a></li>
            """
            
        html_content += """
                    <li><a href="#action-items">Action Items</a></li>
                </ol>
            </div>
        """
        
        # Add Intake Form Validation Section if available
        if excel_validation:
            html_content += """
            <div class="section" id="intake-form-validation">
                <h2>Intake Form Validation</h2>
            """
            
            # Check if the form is valid
            is_valid = excel_validation.get("is_valid", False)
            total_rows = excel_validation.get("total_rows", 0)
            mandatory_fields = excel_validation.get("mandatory_fields", 0)
            unanswered_mandatory = excel_validation.get("unanswered_mandatory", [])
            git_repo_url = excel_validation.get("git_repo_url", "Not provided")
            git_repo_valid = excel_validation.get("git_repo_valid", False)
            
            if is_valid:
                html_content += """
                <div class="validation-status validation-complete">
                    ✅ Intake form is complete. All mandatory fields have been answered.
                </div>
                """
            else:
                html_content += """
                <div class="validation-status validation-incomplete">
                    ❌ Intake form is incomplete. Some mandatory fields have not been answered or Git repository is invalid.
                </div>
                """
                
            html_content += f"""
                <table class="component-table">
                    <tr>
                        <td><strong>Total questions</strong></td>
                        <td>{total_rows}</td>
                    </tr>
                    <tr>
                        <td><strong>Mandatory fields</strong></td>
                        <td>{mandatory_fields}</td>
                    </tr>
                    <tr>
                        <td><strong>Unanswered mandatory fields</strong></td>
                        <td>{len(unanswered_mandatory)}</td>
                    </tr>
                    <tr>
                        <td><strong>Git repository URL</strong></td>
                        <td>{git_repo_url}</td>
                    </tr>
                    <tr>
                        <td><strong>Git repository valid</strong></td>
                        <td>{'✅ Yes' if git_repo_valid else '❌ No'}</td>
                    </tr>
                </table>
            """
            
            if unanswered_mandatory:
                html_content += """
                <h3>Unanswered Mandatory Questions</h3>
                <ul>
                """
                
                for question in unanswered_mandatory:
                    html_content += f"""
                    <li>{question}</li>
                    """
                    
                html_content += """
                </ul>
                """
                
            html_content += """
            </div>
            """
        
        # Component Analysis Section
        html_content += """
        <div class="section" id="component-analysis">
            <h2>Component Analysis</h2>
        """
        
        if excel_components and component_analysis:
            html_content += """
            <p>The following table compares the components declared in the intake form with the components detected in the codebase:</p>
            <table>
                <tr>
                    <th>Component</th>
                    <th>Declared</th>
                    <th>Detected</th>
                    <th>Status</th>
                </tr>
            """
            
            for excel_comp, excel_data in excel_components.items():
                excel_comp_lower = excel_comp.lower()
                excel_declared = excel_data.get("is_yes", False)
                
                # Find if component is in the analysis
                matched = False
                detected = False
                
                for comp_name, comp_data in component_analysis.items():
                    if excel_comp_lower in comp_name.lower() or comp_name.lower() in excel_comp_lower:
                        matched = True
                        detected = comp_data["detected"].lower() == "yes"
                        break
                
                status = "Match" if excel_declared == detected else "Mismatch"
                
                html_content += f"""
                <tr>
                    <td>{excel_comp}</td>
                    <td>{'Yes' if excel_declared else 'No'}</td>
                    <td>{'Yes' if detected else 'No'}</td>
                    <td class="{'match' if status == 'Match' else 'mismatch'}">{status}</td>
                </tr>
                """
                
            html_content += """
            </table>
            """
        elif component_analysis:
            html_content += """
            <p>The following components were detected in the codebase:</p>
            <table>
                <tr>
                    <th>Component</th>
                    <th>Detected</th>
                    <th>Evidence</th>
                </tr>
            """
            
            for comp_name, comp_data in component_analysis.items():
                detected = comp_data.get("detected", "no")
                evidence = comp_data.get("evidence", "None provided")
                
                # Truncate evidence if too long
                if isinstance(evidence, str) and len(evidence) > 50:
                    evidence = evidence[:47] + "..."
                elif isinstance(evidence, list):
                    evidence = ", ".join(evidence[:2])
                    if len(evidence) > 2:
                        evidence += f" (+{len(evidence) - 2} more)"
                
                html_content += f"""
                <tr>
                    <td>{comp_name}</td>
                    <td class="{'component-yes' if detected.lower() == 'yes' else 'component-no'}">{detected}</td>
                    <td>{evidence}</td>
                </tr>
                """
                
            html_content += """
            </table>
            """
        else:
            html_content += """
            <p>No component analysis data available.</p>
            """
        
        html_content += """
        </div>
        """
        
        # Security and Quality Practices Section
        html_content += """
        <div class="section" id="security-and-quality-practices">
            <h2>Security and Quality Practices</h2>
        """
        
        if security_quality_analysis:
            html_content += """
            <p>The following sections summarize the security and quality practices implemented in the codebase:</p>
            """
            
            for category, practices in security_quality_analysis.items():
                # Convert category from snake_case to Title Case
                category_display = category.replace('_', ' ').title()
                html_content += f"""
                <h3>{category_display}</h3>
                <table>
                    <tr>
                        <th>Practice</th>
                        <th>Status</th>
                        <th>Evidence</th>
                    </tr>
                """
                
                for practice_name, practice_data in practices.items():
                    # Convert practice name from snake_case to Title Case
                    practice_display = practice_name.replace('_', ' ').title()
                    
                    status = practice_data.get("implemented", "no")
                    evidence = practice_data.get("evidence", "None provided")
                    
                    status_text = "Implemented" if status == "yes" else "Partially Implemented" if status == "partial" else "Not Implemented"
                    status_class = "component-yes" if status == "yes" else "component-no" if status == "no" else ""
                    
                    # Format evidence
                    if isinstance(evidence, list):
                        evidence = ", ".join(evidence[:2])
                        if len(evidence) > 2:
                            evidence += f" (+{len(evidence) - 2} more)"
                    
                    html_content += f"""
                    <tr>
                        <td>{practice_display}</td>
                        <td class="{status_class}">{status_text}</td>
                        <td>{evidence}</td>
                    </tr>
                    """
                
                html_content += """
                </table>
                """
        else:
            html_content += """
            <p>No security and quality practice analysis data available.</p>
            """
        
        html_content += """
        </div>
        """
        
        # Technology Stack Section
        html_content += """
        <div class="section" id="technology-stack">
            <h2>Technology Stack</h2>
        """
        
        if technology_stack:
            html_content += """
            <p>The following technologies were identified in the codebase:</p>
            """
            
            # Create sections for each technology category
            for category, techs in technology_stack.items():
                if not isinstance(techs, list) or not techs:
                    continue
                    
                # Convert category from snake_case to Title Case
                category_display = category.replace('_', ' ').title()
                html_content += f"""
                <h3>{category_display}</h3>
                <table>
                    <tr>
                        <th>Technology</th>
                        <th>Version</th>
                        <th>Purpose</th>
                        <th>Files</th>
                    </tr>
                """
                
                for tech in techs:
                    if not isinstance(tech, dict):
                        continue
                        
                    tech_name = tech.get("name", "Unknown")
                    tech_version = tech.get("version", "unknown")
                    tech_purpose = tech.get("purpose", "Not specified")
                    tech_files = tech.get("files", [])
                    
                    # Format files list
                    files_display = ", ".join(tech_files[:2])
                    if len(tech_files) > 2:
                        files_display += f" (+{len(tech_files) - 2})"
                    
                    html_content += f"""
                    <tr>
                        <td>{tech_name}</td>
                        <td>{tech_version}</td>
                        <td>{tech_purpose}</td>
                        <td>{files_display}</td>
                    </tr>
                    """
                
                html_content += """
                </table>
                """
        else:
            html_content += """
            <p>No technologies were identified in the codebase.</p>
            """
        
        html_content += """
        </div>
        """
        
        # Detailed Findings Section
        if findings:
            html_content += """
            <div class="section" id="detailed-findings">
                <h2>Detailed Findings</h2>
            """
            
            # Group findings by category
            categories = {}
            for finding in findings:
                category = finding.get("category", "Uncategorized")
                if category not in categories:
                    categories[category] = []
                categories[category].append(finding)
            
            for category, category_findings in categories.items():
                html_content += f"""
                <h3>{category}</h3>
                <div class="finding-list">
                """
                
                for finding in category_findings:
                    severity = finding.get("severity", "Low")
                    description = finding.get("description", "No description provided")
                    recommendation = finding.get("recommendation", "No recommendation provided")
                    location = finding.get("location", {})
                    file_path = location.get("file", "Unknown")
                    line = location.get("line", "Unknown")
                    code = location.get("code", "No code snippet provided")
                    
                    severity_class = f"severity-{severity.lower()}"
                    
                    html_content += f"""
                    <div class="finding {severity_class}">
                        <h4><span class="severity-badge">{severity}</span>{description}</h4>
                        <p><strong>Location:</strong> {file_path}, line {line}</p>
                        <p><strong>Code:</strong> <code>{code}</code></p>
                        <p><strong>Recommendation:</strong> {recommendation}</p>
                    </div>
                    """
                
                html_content += """
                </div>
                """
            
            html_content += """
            </div>
            """
        
        # Add Jira Stories Section if available
        if jira_stories:
            html_content += """
        <div class="section" id="jira-stories">
            <h2>Jira Stories</h2>
        
            <p>The following Jira stories are relevant to this project:</p>
        """
            
            for story in jira_stories:
                html_content += f"""
                <div class="action-item">
                    <h3>{story['key']}: {story['summary']}</h3>
                    <p><strong>Status:</strong> {story['status']}</p>
                    <p><strong>Created:</strong> {story['created']}</p>
                    <p><strong>Last Updated:</strong> {story['updated']}</p>
                """
                
                if story['description']:
                    html_content += f"""
                    <p><strong>Description:</strong></p>
                    <pre>{story['description']}</pre>
                    """
                
                if story['comments']:
                    html_content += """
                    <p><strong>Comments:</strong></p>
                    <ul>
                    """
                    
                    for comment in story['comments']:
                        html_content += f"""
                        <li>
                            <strong>{comment['author']}</strong> ({comment['created']}):
                            <p>{comment['body']}</p>
                        </li>
                        """
                        
                    html_content += """
                    </ul>
                    """
                
                if story['attachments']:
                    html_content += """
                    <p><strong>Attachments:</strong></p>
                    <div class="attachments">
                    """
                    
                    for attachment in story['attachments']:
                        if attachment.get('is_image') and 'local_path' in attachment:
                            html_content += f"""
                            <div class="attachment">
                                <p>{attachment['filename']} ({attachment['size']} bytes)</p>
                                <img src="{attachment['local_path']}" alt="{attachment['filename']}" style="max-width: 100%; max-height: 300px;">
                            </div>
                            """
                        else:
                            html_content += f"""
                            <div class="attachment">
                                <p>{attachment['filename']} ({attachment['size']} bytes)</p>
                            </div>
                            """
                            
                    html_content += """
                    </div>
                    """
                
                html_content += """
                </div>
                """
            
            html_content += """
        </div>
        """

        # Action Items Section
        html_content += """
        <div class="section" id="action-items">
            <h2>Action Items</h2>
            <p>Below are the prioritized action items recommended based on this analysis:</p>
        """
        
        # Define action items
        action_items = []
        
        # 1. Intake form completion if needed
        if excel_validation and not excel_validation.get("is_valid", True):
            action_items.append({
                "priority": "Critical",
                "class": "critical",
                "title": "Complete Intake Form",
                "description": "Complete all mandatory questions in the intake form to ensure accurate project configuration."
            })
            
        # 2. High-priority findings
        high_findings = [f for f in findings if f.get("severity", "Low") == "High"]
        if high_findings:
            action_items.append({
                "priority": "High",
                "class": "high",
                "title": f"Address {len(high_findings)} High-Priority Issues",
                "description": "Fix critical security and quality issues to prevent potential vulnerabilities."
            })
            
        # 3. Component mismatches
        if mismatches_count > 0:
            action_items.append({
                "priority": "Medium",
                "class": "medium",
                "title": "Resolve Component Discrepancies",
                "description": "Reconcile mismatches between declared components and what's detected in the code."
            })
            
        # 4. Missing security practices
        if security_quality_analysis:
            missing_practices_count = 0
            for category, practices in security_quality_analysis.items():
                for practice, details in practices.items():
                    if details.get("implemented", "no") == "no":
                        missing_practices_count += 1
                        
            if missing_practices_count > 0:
                action_items.append({
                    "priority": "Medium",
                    "class": "medium",
                    "title": f"Implement {missing_practices_count} Missing Security Practices",
                    "description": "Address security gaps to improve overall application security posture."
                })
                
        # 5. Medium-priority findings
        medium_findings = [f for f in findings if f.get("severity", "Low") == "Medium"]
        if medium_findings:
            action_items.append({
                "priority": "Medium",
                "class": "medium",
                "title": f"Address {len(medium_findings)} Medium-Priority Issues",
                "description": "Resolve quality and security issues of moderate severity."
            })
        
        # Format action items in the HTML
        if action_items:
            for i, item in enumerate(action_items):
                html_content += f"""
                <div class="action-item {item['class']}">
                    <h3>{i+1}. {item['title']}</h3>
                    <p><strong>Priority:</strong> {item['priority']}</p>
                    <p><strong>Action:</strong> {item['description']}</p>
                </div>
                """
        else:
            html_content += """
            <p>No specific action items identified. The codebase appears to follow good practices.</p>
            """
        
        html_content += """
        </div>
        </body>
        </html>
        """

        # Generate PDF
        pdf_path = None
        try:
            from weasyprint import HTML
            pdf_path = os.path.join(output_dir, "analysis_report.pdf")
            HTML(string=html_content).write_pdf(pdf_path)
            print(f"Generated PDF report: {pdf_path}")
        except Exception as e:
            print(f"Warning: Could not generate PDF: {str(e)}")
            print("Please ensure WeasyPrint is installed correctly.")
        
        # Save both Markdown and HTML
        os.makedirs(output_dir, exist_ok=True)
        
        # Save Markdown
        md_path = os.path.join(output_dir, "analysis_report.md")
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(report)
            
        # Save HTML
        html_path = os.path.join(output_dir, "analysis_report.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
            
        return {
            "markdown": md_path,
            "html": html_path,
            "pdf": pdf_path
        }

    def post(self, shared, prep_res, exec_res):
        # Store the report paths in shared state
        shared["analysis_report"] = exec_res
        
        # Print the report paths
        print(f"\nAnalysis reports generated:")
        print(f"- Markdown: {exec_res['markdown']}")
        print(f"- HTML: {exec_res['html']}")
        if exec_res['pdf']:
            print(f"- PDF: {exec_res['pdf']}")
            
        # Print command to open the HTML report for easier viewing
        print(f"\nOpen the HTML report with: open {exec_res['html']}")
        
        # Print first few components and findings as a sanity check
        print("\nReport Summary:")
        code_analysis = shared.get("code_analysis", {})
        
        # Print component analysis summary
        component_analysis = code_analysis.get("component_analysis", {})
        if component_analysis:
            print("Component Analysis:")
            components = list(component_analysis.keys())
            sample_size = min(5, len(components))
            for i in range(sample_size):
                comp = components[i]
                detected = component_analysis[comp].get("detected", "no")
                print(f"- {comp}: {detected}")
            if len(components) > sample_size:
                print(f"- ... and {len(components) - sample_size} more components")
                
        # Print security implementation summary
        security_quality = code_analysis.get("security_quality_analysis", {})
        if security_quality:
            print("\nSecurity Practice Implementation:")
            for category, practices in security_quality.items():
                implemented = sum(1 for _, details in practices.items() if details.get("implemented") == "yes")
                total = len(practices)
                print(f"- {category.replace('_', ' ').title()}: {implemented}/{total} implemented")
                
        # Print technology stack summary  
        tech_stack = code_analysis.get("technology_stack", {})
        if tech_stack:
            print("\nTechnology Stack:")
            # Only iterate through valid list items in the tech_stack
            for category_name, category_items in tech_stack.items():
                # Add validation before iterating
                if isinstance(category_items, list):
                    print(f"- {category_name.replace('_', ' ').title()}: {len(category_items)} items")
                elif isinstance(category_items, dict):
                    print(f"- {category_name.replace('_', ' ').title()}: {len(category_items)} items")
                else:
                    print(f"- {category_name.replace('_', ' ').title()}: Invalid item type ({type(category_items).__name__})")
        
        # Print findings summary
        findings = code_analysis.get("findings", [])
        if findings:
            severity_counts = {"High": 0, "Medium": 0, "Low": 0}
            for finding in findings:
                severity = finding.get("severity", "Low")
                if severity in severity_counts:
                    severity_counts[severity] += 1
            
            print("\nFindings:")
            print(f"- High Priority: {severity_counts['High']}")
            print(f"- Medium Priority: {severity_counts['Medium']}")
            print(f"- Low Priority: {severity_counts['Low']}")
            
        print("\nCheck the generated reports for complete details.")
        
        # Return a simple string action rather than a complex object
        return "default"


class ProcessExcel(Node):
    def prep(self, shared):
        """
        Prepare input by checking Excel file and sheet name.
        """
        excel_file = shared.get("excel_file")
        if not excel_file:
            print("WARNING: Excel file path not provided in shared state. Skipping Excel processing.")
            return None, None, None
            
        if not os.path.exists(excel_file):
            print(f"WARNING: Excel file not found: {excel_file}. Skipping Excel processing.")
            return None, None, None
            
        sheet_name = shared.get("sheet_name")
        output_dir = shared.get("output_dir", ".")
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        return excel_file, sheet_name, output_dir

    def exec(self, inputs):
        """
        Execute Excel processing logic.
        """
        if inputs is None or len(inputs) != 3 or inputs[0] is None:
            print("No valid Excel file to process.")
            return {
                "component_name": "Unknown Component",
                "repo_url": None,
                "validation": {
                    "total_rows": 0,
                    "mandatory_fields": 0,
                    "git_repo_url": None,
                    "git_repo_valid": False,
                    "is_valid": False,
                    "unanswered_mandatory": [],
                    "component_questions": {}
                }
            }
            
        excel_file, sheet_name, output_dir = inputs
        
        def is_valid_git_repo(url):
            """
            Check if the URL is a valid Git repository.
            """
            if not url:
                return False
            # Common Git URL patterns
            patterns = [
                r'^https?://(?:[\w-]+@)?(?:github\.com|gitlab\.com|bitbucket\.org)/[\w-]+/[\w-]+(?:\.git)?$',
                r'^git@(?:github\.com|gitlab\.com|bitbucket\.org):[\w-]+/[\w-]+(?:\.git)?$'
            ]
            return any(re.match(pattern, url.strip()) for pattern in patterns)

        def find_component_and_repo(sheet):
            """Find component name from first row with value and Git repository URL."""
            component_name = None
            repo_url = None
            git_repo_row = None
            
            # Find component name from first row with value
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                if any(cell.value for cell in row):
                    # Get the first non-empty cell value as component name
                    for cell in row:
                        if cell.value:
                            component_name = str(cell.value).strip()
                            break
                    break
            
            # Find Git repo URL
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                # Check if this row has the Git repo URL in the second column (questions column)
                if len(row) >= 3 and row[1].value and "git repo" in str(row[1].value).lower():
                    git_repo_row = row
                    if row[2].value:  # Answer is in the third column
                        repo_url = str(row[2].value).strip()
                    break
            
            return component_name, repo_url, git_repo_row

        def validate_sheet(sheet):
            """Process the sheet and validate its contents."""
            mandatory_count = 0
            total_rows = 0
            unanswered_mandatory = []
            component_questions = {}
            
            # Get component name, repo URL and its row
            component_name, repo_url, git_repo_row = find_component_and_repo(sheet)
            component_name_row = None
            
            # Validate Git repo URL
            git_repo_valid = is_valid_git_repo(repo_url) if repo_url else False
            
            # Process all rows (questions and answers)
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                # Check if this is the component name row (first row with content)
                if len(row) >= 1 and row[0].value and str(row[0].value).strip() == component_name:
                    component_name_row = row
                    
                if len(row) >= 3 and row[1].value:  # Question in second column
                    question = str(row[1].value).strip()
                    answer = row[2].value
                    
                    total_rows += 1
                    
                    # Check if this is a component question
                    is_component_question = False
                    # Look for patterns like "Is the component using X?" or "Does this use Y?"
                    if any(pattern in question.lower() for pattern in ["is the component using", "does this use", "are you using"]):
                        # Extract component name from question
                        # For example, "Is the component using Venafi?" -> "Venafi"
                        component_name_match = None
                        
                        # Try to extract component name after "using" or "use"
                        for pattern in ["using ", "use "]:
                            if pattern in question.lower():
                                component_parts = question.lower().split(pattern, 1)[1].split("?")[0].strip()
                                if component_parts:
                                    component_name_match = component_parts
                        
                        if component_name_match:
                            # Store the component name and the answer (yes/no)
                            answer_text = str(answer).strip().lower() if answer else ""
                            is_yes = any(yes_word in answer_text for yes_word in ["yes", "y", "true", "1"])
                            component_questions[component_name_match] = {
                                "question": question,
                                "answer": answer_text,
                                "is_yes": is_yes
                            }
                            is_component_question = True
                    
                    # Check if this is a mandatory question
                    # Exclude component questions from mandatory checks
                    is_mandatory = True
                    
                    # Skip component name questions when checking for mandatory fields
                    if is_component_question:
                        is_mandatory = False
                    
                    # Skip component name row and Git repo row
                    if (component_name_row and hasattr(row[0], 'coordinate') and hasattr(component_name_row[0], 'coordinate') and 
                        row[0].coordinate == component_name_row[0].coordinate) or \
                       (git_repo_row and hasattr(row[0], 'coordinate') and hasattr(git_repo_row[0], 'coordinate') and 
                        row[0].coordinate == git_repo_row[0].coordinate):
                        is_mandatory = False
                    
                    if is_mandatory:
                        mandatory_count += 1
                        if not answer or str(answer).strip() == "":
                            unanswered_mandatory.append(question)
            
            # Log unanswered mandatory questions
            if unanswered_mandatory:
                print("\nERROR: Intake form is incomplete. The following mandatory questions are not answered:")
                for question in unanswered_mandatory:
                    print(f"- {question}")
                print("\nPlease complete all mandatory fields before proceeding.")
                print()
            
            return {
                "total_rows": total_rows,
                "mandatory_fields": mandatory_count,
                "git_repo_url": repo_url,
                "git_repo_valid": git_repo_valid,
                "is_valid": mandatory_count > 0 and len(unanswered_mandatory) == 0 and git_repo_valid,
                "unanswered_mandatory": unanswered_mandatory,
                "component_questions": component_questions
            }

        try:
            # Load the Excel file
            wb = load_workbook(excel_file, read_only=True)
            
            # Get the specified sheet or first sheet
            sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            
            # Find component name and repo URL
            component_name, repo_url, _ = find_component_and_repo(sheet)
            
            # Validate the sheet
            validation_results = validate_sheet(sheet)
            
            # Save validation results to JSON
            os.makedirs(output_dir, exist_ok=True)
            validation_file = os.path.join(output_dir, "excel_validation.json")
            
            with open(validation_file, "w") as f:
                json.dump({
                    "component_name": component_name,
                    "repo_url": repo_url,
                    "validation": validation_results
                }, f, indent=2)
            
            return {
                "component_name": component_name or "Unknown Component",
                "repo_url": repo_url,
                "validation": validation_results
            }
            
        except Exception as e:
            print(f"Error processing Excel file: {str(e)}")
            traceback_str = traceback.format_exc()
            print(f"Traceback:\n{traceback_str}")
            
            return {
                "component_name": "Unknown Component",
                "repo_url": None,
                "validation": {
                    "total_rows": 0,
                    "mandatory_fields": 0,
                    "git_repo_url": None,
                    "git_repo_valid": False,
                    "is_valid": False,
                    "unanswered_mandatory": ["Error processing Excel file"],
                    "component_questions": {}
                },
                "error": str(e),
                "traceback": traceback_str
            }

    def post(self, shared, prep_res, exec_res):
        """Store results in shared state."""
        
        # Set default project name if none was found
        if not exec_res.get("component_name") or exec_res.get("component_name") == "Unknown Component":
            project_name = shared.get("project_name", "Unknown Component")
            exec_res["component_name"] = project_name
        
        # Handle validation result display
        validation = exec_res.get("validation", {})
        is_valid = validation.get("is_valid", False)
        unanswered_mandatory = validation.get("unanswered_mandatory", [])
        
        if not is_valid:
            if unanswered_mandatory:
                print("Intake form is incomplete. Please complete all mandatory fields before proceeding.")
            else:
                print("Excel validation failed. Please check the excel_validation.json file for details.")
            
        # Store the Git repo URL in shared state
        shared["repo_url"] = exec_res.get("repo_url")
        
        # Store the component name as project name
        shared["project_name"] = exec_res.get("component_name")
        
        # Store validation results
        shared["excel_validation"] = exec_res.get("validation", {})
        
        # Check if we have either a valid repo URL or a local directory to proceed
        if (shared.get("repo_url") or shared.get("local_dir")) and not exec_res.get("error"):
            print(f"Excel processing successful. Component: {shared['project_name']}")
            print(f"Repository URL: {shared.get('repo_url', 'Not specified')}")
            print(f"Local directory: {shared.get('local_dir', 'Not specified')}")
            return "success"
        else:
            print("Excel processing completed, but no repository URL or local directory available.")
            print("Cannot proceed with code analysis. Please check the Excel file and try again.")
            return "error"


class FetchJiraStories(Node):
    def prep(self, shared):
        print("\nDEBUG: Starting FetchJiraStories prep")
        
        # Get Jira credentials and configuration from environment variables
        # or let users pass them as command line arguments
        jira_url = os.environ.get("JIRA_URL", shared.get("jira_url"))
        jira_username = os.environ.get("JIRA_USERNAME", shared.get("jira_username"))
        jira_api_token = os.environ.get("JIRA_API_TOKEN", shared.get("jira_api_token"))
        jira_project_key = os.environ.get("JIRA_PROJECT_KEY", shared.get("jira_project_key", "XYZ"))
        output_dir = shared.get("output_dir", "analysis_output")
        
        # DEBUG FLAG - Set to True to use sample data instead of real Jira
        use_sample_data = True
        
        if use_sample_data:
            print("Using sample Jira data for testing")
            return {"use_sample": True, "output_dir": output_dir, "project_key": jira_project_key}
            
        if not jira_url or not jira_username or not jira_api_token:
            print("Warning: Jira credentials not found. Skipping Jira integration.")
            return None
            
        return {"use_sample": False, "jira_url": jira_url, "jira_username": jira_username, 
                "jira_api_token": jira_api_token, "project_key": jira_project_key, "output_dir": output_dir}
        
    def exec(self, prep_res):
        if not prep_res:
            # Skip Jira integration if credentials aren't available
            print("Skipping Jira integration due to missing credentials")
            return None
            
        # Check if we should use sample data
        if prep_res.get("use_sample", False):
            return self._generate_sample_stories(prep_res["project_key"], prep_res["output_dir"])
            
        jira_url = prep_res["jira_url"]
        jira_username = prep_res["jira_username"]
        jira_api_token = prep_res["jira_api_token"]
        jira_project_key = prep_res["project_key"]
        output_dir = prep_res["output_dir"]
        
        try:
            from jira import JIRA
            
            print(f"Connecting to Jira at {jira_url} with username {jira_username}")
            print(f"Using project key: {jira_project_key}")
            
            # Connect to Jira
            try:
                jira = JIRA(
                    server=jira_url,
                    basic_auth=(jira_username, jira_api_token)
                )
                
                # Verify connection by trying to access the server info
                server_info = jira.server_info()
                print(f"Successfully connected to Jira server version: {server_info.get('version', 'unknown')}")
            except Exception as conn_error:
                print(f"Error connecting to Jira: {str(conn_error)}")
                return None
            
            # Search for issues that start with the specified project key
            try:
                jql_query = f'project = "{jira_project_key}" ORDER BY updated DESC'
                print(f"Executing JQL query: {jql_query}")
                issues = jira.search_issues(jql_query, maxResults=10)
                print(f"Found {len(issues)} Jira issues")
            except Exception as query_error:
                print(f"Error searching for Jira issues: {str(query_error)}")
                return None
            
            if not issues:
                print(f"No issues found for project key: {jira_project_key}")
                return []
                
            # Extract relevant information from each issue
            stories = []
            for issue in issues:
                print(f"Processing issue: {issue.key}")
                # Get issue details
                story = {
                    'key': issue.key,
                    'summary': issue.fields.summary,
                    'status': issue.fields.status.name,
                    'description': issue.fields.description or '',
                    'created': issue.fields.created,
                    'updated': issue.fields.updated,
                    'comments': [],
                    'attachments': []
                }
                
                # Get comments
                if hasattr(issue.fields, 'comment'):
                    for comment in issue.fields.comment.comments:
                        story['comments'].append({
                            'author': comment.author.displayName,
                            'body': comment.body,
                            'created': comment.created
                        })
                
                # Get attachments - especially images
                if hasattr(issue.fields, 'attachment'):
                    for attachment in issue.fields.attachment:
                        is_image = attachment.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))
                        
                        # For images, save them locally
                        attachment_data = {
                            'filename': attachment.filename,
                            'created': attachment.created,
                            'is_image': is_image,
                            'content_type': attachment.mimeType,
                            'size': attachment.size
                        }
                        
                        if is_image:
                            try:
                                # Download image and save it
                                image_data = jira.download_attachment(attachment.id)
                                attachments_dir = os.path.join(output_dir, 'jira_attachments')
                                os.makedirs(attachments_dir, exist_ok=True)
                                
                                image_path = os.path.join(attachments_dir, attachment.filename)
                                with open(image_path, 'wb') as f:
                                    f.write(image_data)
                                    
                                # Store relative path for embedding in HTML
                                # Use a relative path that works in HTML context
                                attachment_data['local_path'] = f"jira_attachments/{attachment.filename}"
                                print(f"Saved attachment: {image_path}")
                            except Exception as e:
                                print(f"Error downloading attachment {attachment.filename}: {str(e)}")
                        
                        story['attachments'].append(attachment_data)
                
                stories.append(story)
            
            return stories
            
        except ImportError:
            print("Error: Jira package not installed. Please install it using 'pip install jira'.")
            return None
        except Exception as e:
            print(f"Error fetching Jira stories: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def _generate_sample_stories(self, project_key, output_dir):
        """Generate sample Jira stories for testing"""
        print("Generating sample Jira stories for testing")
        
        # Create sample stories
        stories = []
        
        # Story 1 - Feature request with comments and attachments
        story1 = {
            'key': f'{project_key}-101',
            'summary': 'Implement Jira integration for code analyzer',
            'status': 'In Progress',
            'description': 'As a user, I want to see my relevant Jira stories in the code analysis report so that I can track issues related to my codebase.\n\n**Acceptance Criteria:**\n- Fetch stories from Jira API\n- Display summary, status, and description\n- Show comments and attachments\n- Integrate with existing report',
            'created': '2023-05-15T10:30:00.000+0000',
            'updated': '2023-05-16T14:45:00.000+0000',
            'comments': [
                {
                    'author': 'John Developer',
                    'body': 'I\'ve started working on this. Initial API connection is working.',
                    'created': '2023-05-15T15:22:00.000+0000'
                },
                {
                    'author': 'Maria Manager',
                    'body': 'Great! Make sure to handle authentication errors gracefully.',
                    'created': '2023-05-15T16:05:00.000+0000'
                }
            ],
            'attachments': []
        }
        
        # Create a sample image attachment
        try:
            # Create a simple image
            from PIL import Image, ImageDraw, ImageFont
            import random
            
            attachments_dir = os.path.join(output_dir, 'jira_attachments')
            os.makedirs(attachments_dir, exist_ok=True)
            
            # Create a colored rectangle with text
            img = Image.new('RGB', (400, 200), color=(73, 109, 137))
            d = ImageDraw.Draw(img)
            d.text((10, 10), f"Sample Jira Image\n{project_key}-101", fill=(255, 255, 0))
            d.rectangle([(50, 50), (350, 150)], outline=(255, 255, 255))
            
            image_filename = f"{project_key}_design_mockup.png"
            image_path = os.path.join(attachments_dir, image_filename)
            
            # Save the image
            img.save(image_path)
            
            # Add to story attachments
            story1['attachments'].append({
                'filename': image_filename,
                'created': '2023-05-15T12:00:00.000+0000',
                'is_image': True,
                'content_type': 'image/png',
                'size': os.path.getsize(image_path),
                'local_path': f"jira_attachments/{image_filename}"
            })
            
            print(f"Created sample image attachment: {image_path}")
            
        except ImportError:
            print("PIL library not available, skipping image generation")
        except Exception as e:
            print(f"Error creating sample image: {str(e)}")
        
        # Story 2 - Bug report
        story2 = {
            'key': f'{project_key}-102',
            'summary': 'Fix broken CSS in report generation',
            'status': 'To Do',
            'description': 'The CSS styles in the generated report are not being applied correctly. The Jira stories section is not displaying images properly.\n\n**Steps to reproduce:**\n1. Generate a report with Jira stories\n2. Open the HTML report\n3. Notice the images are not displayed',
            'created': '2023-05-17T09:15:00.000+0000',
            'updated': '2023-05-17T09:15:00.000+0000',
            'comments': [],
            'attachments': []
        }
        
        # Story 3 - Completed task
        story3 = {
            'key': f'{project_key}-100',
            'summary': 'Set up project structure and dependencies',
            'status': 'Done',
            'description': 'Create the initial project structure with requirements.txt, main.py, and necessary modules.',
            'created': '2023-05-10T08:00:00.000+0000',
            'updated': '2023-05-12T16:30:00.000+0000',
            'comments': [
                {
                    'author': 'John Developer',
                    'body': 'Completed and ready for review.',
                    'created': '2023-05-12T15:45:00.000+0000'
                },
                {
                    'author': 'Sarah Reviewer',
                    'body': 'Looks good! Approved.',
                    'created': '2023-05-12T16:30:00.000+0000'
                }
            ],
            'attachments': []
        }
        
        # Add the stories to the list
        stories.extend([story1, story2, story3])
        
        print(f"Generated {len(stories)} sample Jira stories")
        return stories
    
    def post(self, shared, prep_res, exec_res):
        if exec_res:
            print(f"\nFetched {len(exec_res)} Jira stories")
            shared["jira_stories"] = exec_res
        else:
            print("No Jira stories fetched or error occurred")
            shared["jira_stories"] = []
            
        return "default"
