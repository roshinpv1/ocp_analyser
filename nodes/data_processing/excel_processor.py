import os
import re
import json
import traceback
from openpyxl import load_workbook
from core.genflow import Node

class ProcessExcel(Node):
    def prep(self, shared):
        """
        Prepare input by checking Excel file and sheet name.
        """
        excel_file = shared.get("excel_file")
        if not excel_file:
            print("WARNING: Excel file path not provided in shared state. Skipping Excel processing.")
            return None, None, None
            
        if not os.path.exists(excel_file):
            print(f"WARNING: Excel file not found: {excel_file}. Skipping Excel processing.")
            return None, None, None
            
        sheet_name = shared.get("sheet_name")
        output_dir = shared.get("output_dir", ".")
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        return excel_file, sheet_name, output_dir

    def exec(self, inputs):
        """
        Execute Excel processing logic.
        """
        if inputs is None or len(inputs) != 3 or inputs[0] is None:
            print("No valid Excel file to process.")
            return {
                "component_name": "Unknown Component",
                "repo_url": None,
                "validation": {
                    "total_rows": 0,
                    "mandatory_fields": 0,
                    "git_repo_url": None,
                    "git_repo_valid": False,
                    "is_valid": False,
                    "unanswered_mandatory": [],
                    "component_questions": {}
                }
            }
            
        excel_file, sheet_name, output_dir = inputs
        
        def is_valid_git_repo(url):
            """
            Check if the URL is a valid Git repository.
            """
            if not url:
                return False
            # Common Git URL patterns
            patterns = [
                r'^https?://(?:[\w-]+@)?(?:github\.com|gitlab\.com|bitbucket\.org)/[\w-]+/[\w-]+(?:\.git)?$',
                r'^git@(?:github\.com|gitlab\.com|bitbucket\.org):[\w-]+/[\w-]+(?:\.git)?$'
            ]
            return any(re.match(pattern, url.strip()) for pattern in patterns)

        def find_component_and_repo(sheet):
            """Find component name from first row with value and Git repository URL."""
            component_name = None
            repo_url = None
            git_repo_row = None
            
            # Find component name from first row with value
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                if any(cell.value for cell in row):
                    # Get the first non-empty cell value as component name
                    for cell in row:
                        if cell.value:
                            component_name = str(cell.value).strip()
                            break
                    break
            
            # Find Git repo URL - look for rows with "git repo" text
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                # Check for Git repo keywords in any cell
                for i, cell in enumerate(row):
                    if cell and cell.value and isinstance(cell.value, str) and "git repo" in cell.value.lower():
                        git_repo_row = row
                        
                        # Try to find URL in this row - check other cells first
                        for j, value_cell in enumerate(row):
                            if j != i and value_cell and value_cell.value and isinstance(value_cell.value, str) and ("http" in value_cell.value.lower() or "git@" in value_cell.value.lower()):
                                repo_url = str(value_cell.value).strip()
                                break
                        
                        # If no URL found in other cells, check if the keyword cell itself contains a URL
                        if not repo_url and ("http" in cell.value.lower() or "git@" in cell.value.lower()):
                            # Extract URL from text if it contains both the keyword and URL
                            url_match = re.search(r'(https?://\S+|git@\S+)', cell.value)
                            if url_match:
                                repo_url = url_match.group(0).strip()
                        
                        if repo_url:
                            break
                
                if git_repo_row and repo_url:
                    break
            
            return component_name, repo_url, git_repo_row

        def validate_sheet(sheet):
            """Process the sheet and validate its contents."""
            mandatory_count = 0
            total_rows = 0
            unanswered_mandatory = []
            component_questions = {}
            
            # Get component name, repo URL and its row
            component_name, repo_url, git_repo_row = find_component_and_repo(sheet)
            component_name_row = None
            
            # Validate Git repo URL
            git_repo_valid = is_valid_git_repo(repo_url) if repo_url else False
            
            # Process all rows (questions and answers)
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                # Check if this is the component name row (first row with content)
                if len(row) >= 1 and row[0].value and str(row[0].value).strip() == component_name:
                    component_name_row = row
                
                # Skip empty rows
                if not any(cell.value for cell in row):
                    continue
                
                # Extract question and answer from the row
                question = None
                answer = None
                
                # Check all cells in the row to identify question and answer
                for i, cell in enumerate(row):
                    if not cell.value:
                        continue
                        
                    cell_text = str(cell.value).strip()
                    
                    # If we don't have a question yet, and this looks like a question, set it as the question
                    if not question and any(q_word in cell_text.lower() for q_word in ["is ", "are ", "does ", "do ", "how ", "what ", "when ", "where ", "why ", "which "]):
                        question = cell_text
                    # Otherwise, if we don't have an answer yet, this might be an answer
                    elif not answer:
                        answer = cell_text
                    # If we already have both, and this cell has a URL, it might be a better answer
                    elif answer and ("http" in cell_text.lower() or "git@" in cell_text.lower()):
                        answer = cell_text
                
                # If we only found one cell with content, treat it as an answer with a default question
                if not question and answer:
                    question = "Extracted Value"
                
                # Skip rows where we couldn't identify a question-answer pair
                if not question:
                    continue
                    
                total_rows += 1
                
                # Check if this is a component question
                is_component_question = False
                # Look for patterns like "Is the component using X?" or "Does this use Y?"
                if any(pattern in question.lower() for pattern in ["is the component using", "does this use", "are you using"]):
                    # Extract component name from question
                    # For example, "Is the component using Venafi?" -> "Venafi"
                    component_name_match = None
                    
                    # Try to extract component name after "using" or "use"
                    for pattern in ["using ", "use "]:
                        if pattern in question.lower():
                            component_parts = question.lower().split(pattern, 1)[1].split("?")[0].strip()
                            if component_parts:
                                component_name_match = component_parts
                    
                    if component_name_match:
                        # Store the component name and the answer (yes/no)
                        answer_text = str(answer).strip().lower() if answer else ""
                        is_yes = any(yes_word in answer_text for yes_word in ["yes", "y", "true", "1"])
                        component_questions[component_name_match] = {
                            "question": question,
                            "answer": answer_text,
                            "is_yes": is_yes
                        }
                        is_component_question = True
                
                # Check if this is a mandatory question
                # Exclude component questions from mandatory checks
                is_mandatory = True
                
                # Skip component questions when checking for mandatory fields
                if is_component_question:
                    is_mandatory = False
                
                # Skip component name row and Git repo row
                if (component_name_row and hasattr(row[0], 'coordinate') and hasattr(component_name_row[0], 'coordinate') and 
                    row[0].coordinate == component_name_row[0].coordinate) or \
                   (git_repo_row and hasattr(row[0], 'coordinate') and hasattr(git_repo_row[0], 'coordinate') and 
                    row[0].coordinate == git_repo_row[0].coordinate):
                    is_mandatory = False
                
                if is_mandatory:
                    mandatory_count += 1
                    if not answer or str(answer).strip() == "":
                        unanswered_mandatory.append(question)
            
            # Log unanswered mandatory questions
            if unanswered_mandatory:
                print("\nERROR: Intake form is incomplete. The following mandatory questions are not answered:")
                for question in unanswered_mandatory:
                    print(f"- {question}")
                print("\nPlease complete all mandatory fields before proceeding.")
                print()
            
            return {
                "total_rows": total_rows,
                "mandatory_fields": mandatory_count,
                "git_repo_url": repo_url,
                "git_repo_valid": git_repo_valid,
                "is_valid": mandatory_count > 0 and len(unanswered_mandatory) == 0 and git_repo_valid,
                "unanswered_mandatory": unanswered_mandatory,
                "component_questions": component_questions
            }

        try:
            # Load the Excel file
            wb = load_workbook(excel_file, read_only=True)
            
            # Get the specified sheet or first sheet
            sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            
            # Find component name and repo URL
            component_name, repo_url, _ = find_component_and_repo(sheet)
            
            # Validate the sheet
            validation_results = validate_sheet(sheet)
            
            # Save validation results to JSON
            os.makedirs(output_dir, exist_ok=True)
            validation_file = os.path.join(output_dir, "excel_validation.json")
            
            with open(validation_file, "w") as f:
                json.dump({
                    "component_name": component_name,
                    "repo_url": repo_url,
                    "validation": validation_results
                }, f, indent=2)
            
            return {
                "component_name": component_name or "Unknown Component",
                "repo_url": repo_url,
                "validation": validation_results
            }
            
        except Exception as e:
            print(f"Error processing Excel file: {str(e)}")
            traceback_str = traceback.format_exc()
            print(f"Traceback:\n{traceback_str}")
            
            return {
                "component_name": "Unknown Component",
                "repo_url": None,
                "validation": {
                    "total_rows": 0,
                    "mandatory_fields": 0,
                    "git_repo_url": None,
                    "git_repo_valid": False,
                    "is_valid": False,
                    "unanswered_mandatory": ["Error processing Excel file"],
                    "component_questions": {}
                },
                "error": str(e),
                "traceback": traceback_str
            }

    def post(self, shared, prep_res, exec_res):
        """Store results in shared state."""
        
        # Set default project name if none was found
        if not exec_res.get("component_name") or exec_res.get("component_name") == "Unknown Component":
            project_name = shared.get("project_name", "Unknown Component")
            exec_res["component_name"] = project_name
        
        # Handle validation result display
        validation = exec_res.get("validation", {})
        is_valid = validation.get("is_valid", False)
        unanswered_mandatory = validation.get("unanswered_mandatory", [])
        
        if not is_valid:
            if unanswered_mandatory:
                print("Intake form is incomplete. Please complete all mandatory fields before proceeding.")
            else:
                print("Excel validation failed. Please check the excel_validation.json file for details.")
            
        # Store the Git repo URL in shared state
        shared["repo_url"] = exec_res.get("repo_url")
        
        # Store the component name as project name
        shared["project_name"] = exec_res.get("component_name")
        
        # Store validation results
        shared["excel_validation"] = exec_res.get("validation", {})
        
        # Check if we have either a valid repo URL or a local directory to proceed
        if (shared.get("repo_url") or shared.get("local_dir")) and not exec_res.get("error"):
            print(f"Excel processing successful. Component: {shared['project_name']}")
            print(f"Repository URL: {shared.get('repo_url', 'Not specified')}")
            print(f"Local directory: {shared.get('local_dir', 'Not specified')}")
            return "success"
        else:
            print("Excel processing completed, but no repository URL or local directory available.")
            print("Cannot proceed with code analysis. Please check the Excel file and try again.")
            return "error" 